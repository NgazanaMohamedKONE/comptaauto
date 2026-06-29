from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.db.models import Count, Sum, Q, Avg
from django.utils import timezone
from datetime import timedelta, datetime
from django.contrib.auth import login as auth_login, get_user_model
from accounts.models import User, Entreprise
from comptabilite.models import Ecriture
from ocr_app.models import FactureOCR
from communication.models import Ticket, Conversation, DemandeRappel
from abonnements.models import Abonnement, Paiement, Forfait
from .models import LogAudit, NoteEntreprise, TagEntreprise, EntrepriseTag, SessionImpersonate
from .utils import log_action

User = get_user_model()


def admin_required(view_func):
    """Decorator : verifie que c'est un super admin"""
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('login')
        if not request.user.is_super_admin:
            return redirect('dashboard')
        return view_func(request, *args, **kwargs)
    return wrapper


# ============================================
# DASHBOARD ADMIN
# ============================================
@admin_required
def dashboard_admin(request):
    """Dashboard admin enrichi"""
    now = timezone.now()
    debut_mois = now.replace(day=1, hour=0, minute=0, second=0)
    debut_annee = now.replace(month=1, day=1, hour=0, minute=0, second=0)

    # Stats globales
    total_entreprises = Entreprise.objects.count()
    total_users = User.objects.filter(role='ENTREPRISE').count()
    entreprises_actives = Entreprise.objects.filter(statut='ACTIF').count()
    abonnements_actifs = Abonnement.objects.filter(statut='ACTIF').count()

    # Stats financiers
    revenus_total = Paiement.objects.filter(statut='VALIDE').aggregate(s=Sum('montant'))['s'] or 0
    revenus_mois = Paiement.objects.filter(statut='VALIDE', created_at__gte=debut_mois).aggregate(s=Sum('montant'))['s'] or 0
    revenus_annee = Paiement.objects.filter(statut='VALIDE', created_at__gte=debut_annee).aggregate(s=Sum('montant'))['s'] or 0

    # MRR (Monthly Recurring Revenue)
    mrr = 0
    for abo in Abonnement.objects.filter(statut='ACTIF'):
        if abo.forfait.duree_jours <= 31:
            mrr += float(abo.forfait.prix)
        else:
            mrr += float(abo.forfait.prix) * 30 / abo.forfait.duree_jours

    # Alertes
    paiements_en_attente = Paiement.objects.filter(statut='EN_ATTENTE').count()
    tickets_ouverts = Ticket.objects.filter(statut__in=['OUVERT', 'EN_COURS']).count()
    rappels_attente = DemandeRappel.objects.filter(statut='EN_ATTENTE').count()
    abonnements_expirent_bientot = Abonnement.objects.filter(
        statut='ACTIF', date_fin__lte=now + timedelta(days=7)
    ).count()

    # Top entreprises (par nb ecritures)
    top_entreprises = Entreprise.objects.annotate(
        nb_ecr=Count('ecritures')
    ).order_by('-nb_ecr')[:5]

    # Nouvelles entreprises ce mois
    nouvelles_entreprises = Entreprise.objects.filter(date_inscription__gte=debut_mois).count()

    # Repartition par forfait (donut)
    repartition = []
    for f in Forfait.objects.filter(actif=True):
        count = Abonnement.objects.filter(forfait=f, statut='ACTIF').count()
        if count > 0:
            repartition.append({'nom': f.nom, 'count': count, 'couleur': f.couleur})

    # Activite recente
    logs_recents = LogAudit.objects.all()[:10]

    return render(request, 'admin_panel/dashboard.html', {
        'total_entreprises': total_entreprises,
        'total_users': total_users,
        'entreprises_actives': entreprises_actives,
        'abonnements_actifs': abonnements_actifs,
        'revenus_total': revenus_total,
        'revenus_mois': revenus_mois,
        'revenus_annee': revenus_annee,
        'mrr': int(mrr),
        'arr': int(mrr * 12),
        'paiements_en_attente': paiements_en_attente,
        'tickets_ouverts': tickets_ouverts,
        'rappels_attente': rappels_attente,
        'abonnements_expirent_bientot': abonnements_expirent_bientot,
        'top_entreprises': top_entreprises,
        'nouvelles_entreprises': nouvelles_entreprises,
        'repartition': repartition,
        'logs_recents': logs_recents,
    })


@admin_required
def stats_data(request):
    """API : donnees pour graphiques"""
    now = timezone.now()

    # Revenus 12 derniers mois
    mois_labels = []
    mois_revenus = []
    for i in range(11, -1, -1):
        debut = (now.replace(day=1) - timedelta(days=i*30)).replace(day=1)
        fin = (debut + timedelta(days=32)).replace(day=1)
        revenus = Paiement.objects.filter(
            statut='VALIDE', created_at__gte=debut, created_at__lt=fin
        ).aggregate(s=Sum('montant'))['s'] or 0
        mois_labels.append(debut.strftime('%b %Y'))
        mois_revenus.append(float(revenus))

    # Nouveaux clients 12 derniers mois
    mois_clients = []
    for i in range(11, -1, -1):
        debut = (now.replace(day=1) - timedelta(days=i*30)).replace(day=1)
        fin = (debut + timedelta(days=32)).replace(day=1)
        count = Entreprise.objects.filter(date_inscription__gte=debut, date_inscription__lt=fin).count()
        mois_clients.append(count)

    return JsonResponse({
        'revenus': {'labels': mois_labels, 'data': mois_revenus},
        'clients': {'labels': mois_labels, 'data': mois_clients},
    })


# ============================================
# GESTION UTILISATEURS
# ============================================
@admin_required
def users_list(request):
    """Liste de tous les utilisateurs"""
    q = request.GET.get('q', '')
    users = User.objects.all().order_by('-date_joined')
    if q:
        users = users.filter(Q(email__icontains=q) | Q(username__icontains=q) | Q(first_name__icontains=q) | Q(last_name__icontains=q))
    return render(request, 'admin_panel/users_list.html', {'users': users, 'q': q})


@admin_required
def user_detail(request, pk):
    """Detail d'un utilisateur"""
    user = get_object_or_404(User, pk=pk)
    logs = LogAudit.objects.filter(user=user)[:20]
    entreprise = getattr(user, 'entreprise', None)
    return render(request, 'admin_panel/user_detail.html', {
        'user_obj': user, 'logs': logs, 'entreprise': entreprise,
    })


@admin_required
def user_create(request):
    """Creer un utilisateur"""
    if request.method == 'POST':
        try:
            user = User.objects.create_user(
                email=request.POST.get('email'),
                username=request.POST.get('username'),
                password=request.POST.get('password'),
                first_name=request.POST.get('first_name', ''),
                last_name=request.POST.get('last_name', ''),
                telephone=request.POST.get('telephone', ''),
                role=request.POST.get('role', 'ENTREPRISE'),
            )
            if request.POST.get('role') == 'SUPER_ADMIN':
                user.is_staff = True
                user.is_superuser = True
                user.save()
            log_action(request, 'CREATION', f"Creation utilisateur {user.email}", user)
            messages.success(request, f"Utilisateur {user.email} cree")
            return redirect('admin_users_list')
        except Exception as e:
            messages.error(request, f"Erreur: {e}")
    return render(request, 'admin_panel/user_form.html', {'user_obj': None})


@admin_required
def user_edit(request, pk):
    """Modifier un utilisateur"""
    user = get_object_or_404(User, pk=pk)
    if request.method == 'POST':
        user.email = request.POST.get('email', user.email)
        user.username = request.POST.get('username', user.username)
        user.first_name = request.POST.get('first_name', user.first_name)
        user.last_name = request.POST.get('last_name', user.last_name)
        user.telephone = request.POST.get('telephone', user.telephone)
        user.role = request.POST.get('role', user.role)
        user.is_active = request.POST.get('is_active') == 'on'
        user.save()
        log_action(request, 'MODIFICATION', f"Modification utilisateur {user.email}", user)
        messages.success(request, "Utilisateur modifie")
        return redirect('admin_user_detail', pk=user.pk)
    return render(request, 'admin_panel/user_form.html', {'user_obj': user})


@admin_required
def user_delete(request, pk):
    """Supprimer un utilisateur"""
    user = get_object_or_404(User, pk=pk)
    if user == request.user:
        return JsonResponse({'error': 'Vous ne pouvez pas vous supprimer'}, status=400)
    email = user.email
    user.delete()
    log_action(request, 'SUPPRESSION', f"Suppression utilisateur {email}")
    return JsonResponse({'status': 'ok'})


@admin_required
def user_reset_password(request, pk):
    """Reinitialiser mot de passe"""
    user = get_object_or_404(User, pk=pk)
    new_password = request.POST.get('password', 'Temp2026!')
    user.set_password(new_password)
    user.save()
    log_action(request, 'MODIFICATION', f"Reset password {user.email}", user)

    # Notif
    from communication.models import Notification
    Notification.objects.create(
        user=user, type_notif='SYSTEME',
        titre='Mot de passe reinitialise',
        message=f'Votre nouveau mot de passe : {new_password}. Changez-le rapidement !',
    )
    return JsonResponse({'status': 'ok', 'password': new_password})


@admin_required
def user_toggle_active(request, pk):
    """Activer/desactiver un utilisateur"""
    user = get_object_or_404(User, pk=pk)
    user.is_active = not user.is_active
    user.save()
    log_action(request, 'MODIFICATION', f"{'Active' if user.is_active else 'Desactive'} {user.email}", user)
    return JsonResponse({'status': 'ok', 'is_active': user.is_active})


@admin_required
def user_impersonate(request, pk):
    """Se connecter comme un autre utilisateur"""
    user = get_object_or_404(User, pk=pk)
    if user.is_super_admin:
        messages.error(request, "Impossible d'impersonifier un autre super admin")
        return redirect('admin_user_detail', pk=user.pk)

    SessionImpersonate.objects.create(
        admin=request.user, user_cible=user,
        raison=request.POST.get('raison', 'Debug / Support'),
    )
    log_action(request, 'IMPERSONATE', f"Impersonate de {user.email}", user)

    # Sauvegarder l'admin original en session
    request.session['admin_original_id'] = request.user.id
    auth_login(request, user, backend='django.contrib.auth.backends.ModelBackend')
    messages.warning(request, f"Vous etes connecte comme {user.email}. Cliquez sur 'Revenir admin' pour quitter.")
    return redirect('dashboard')


def stop_impersonate(request):
    """Revenir au compte admin original"""
    admin_id = request.session.get('admin_original_id')
    if admin_id:
        admin = User.objects.get(pk=admin_id)
        auth_login(request, admin, backend='django.contrib.auth.backends.ModelBackend')
        del request.session['admin_original_id']
        messages.success(request, "Retour au compte admin")
    return redirect('admin_dashboard')


# ============================================
# GESTION ENTREPRISES AVANCEE
# ============================================
@admin_required
def entreprises_list(request):
    """Liste enrichie des entreprises"""
    q = request.GET.get('q', '')
    statut_filter = request.GET.get('statut', '')
    forfait_filter = request.GET.get('forfait', '')

    entreprises = Entreprise.objects.all().select_related('responsable').prefetch_related('tags__tag', 'abonnements')

    if q:
        entreprises = entreprises.filter(Q(nom__icontains=q) | Q(email_contact__icontains=q) | Q(secteur__icontains=q))
    if statut_filter:
        entreprises = entreprises.filter(statut=statut_filter)

    tags = TagEntreprise.objects.all()

    return render(request, 'admin_panel/entreprises_list.html', {
        'entreprises': entreprises, 'q': q, 'tags': tags,
        'statut_filter': statut_filter,
    })


@admin_required
def entreprise_detail(request, pk):
    """Detail complet d'une entreprise"""
    entreprise = get_object_or_404(Entreprise, pk=pk)

    abonnement_actuel = Abonnement.objects.filter(entreprise=entreprise, statut='ACTIF').first()
    historique_abos = Abonnement.objects.filter(entreprise=entreprise)
    paiements = Paiement.objects.filter(entreprise=entreprise)
    nb_ecritures = Ecriture.objects.filter(entreprise=entreprise).count()
    nb_ocr = FactureOCR.objects.filter(entreprise=entreprise).count()
    tickets = Ticket.objects.filter(entreprise=entreprise)[:10]
    notes = NoteEntreprise.objects.filter(entreprise=entreprise)
    tags = EntrepriseTag.objects.filter(entreprise=entreprise).select_related('tag')
    tags_disponibles = TagEntreprise.objects.exclude(
        id__in=tags.values_list('tag_id', flat=True)
    )

    revenus_entreprise = paiements.filter(statut='VALIDE').aggregate(s=Sum('montant'))['s'] or 0

    return render(request, 'admin_panel/entreprise_detail.html', {
        'entreprise': entreprise,
        'abonnement_actuel': abonnement_actuel,
        'historique_abos': historique_abos,
        'paiements': paiements,
        'nb_ecritures': nb_ecritures,
        'nb_ocr': nb_ocr,
        'tickets': tickets,
        'notes': notes,
        'tags': tags,
        'tags_disponibles': tags_disponibles,
        'revenus_entreprise': revenus_entreprise,
    })


@admin_required
def entreprise_edit(request, pk):
    entreprise = get_object_or_404(Entreprise, pk=pk)
    if request.method == 'POST':
        entreprise.nom = request.POST.get('nom', entreprise.nom)
        entreprise.secteur = request.POST.get('secteur', entreprise.secteur)
        entreprise.adresse = request.POST.get('adresse', entreprise.adresse)
        entreprise.telephone = request.POST.get('telephone', entreprise.telephone)
        entreprise.email_contact = request.POST.get('email_contact', entreprise.email_contact)
        entreprise.formule = request.POST.get('formule', entreprise.formule)
        entreprise.statut = request.POST.get('statut', entreprise.statut)
        entreprise.save()
        log_action(request, 'MODIFICATION', f"Modification entreprise {entreprise.nom}", entreprise)
        messages.success(request, "Entreprise modifiee")
        return redirect('admin_entreprise_detail', pk=entreprise.pk)
    return render(request, 'admin_panel/entreprise_form.html', {'entreprise': entreprise})


@admin_required
def ajouter_note(request, pk):
    entreprise = get_object_or_404(Entreprise, pk=pk)
    if request.method == 'POST':
        NoteEntreprise.objects.create(
            entreprise=entreprise, auteur=request.user,
            contenu=request.POST.get('contenu'),
            importante=request.POST.get('importante') == 'on',
        )
        messages.success(request, "Note ajoutee")
    return redirect('admin_entreprise_detail', pk=pk)


@admin_required
def supprimer_note(request, pk):
    note = get_object_or_404(NoteEntreprise, pk=pk)
    entreprise_pk = note.entreprise.pk
    note.delete()
    return redirect('admin_entreprise_detail', pk=entreprise_pk)


@admin_required
def ajouter_tag_entreprise(request, pk):
    entreprise = get_object_or_404(Entreprise, pk=pk)
    tag_id = request.POST.get('tag_id')
    if tag_id:
        tag = get_object_or_404(TagEntreprise, pk=tag_id)
        EntrepriseTag.objects.get_or_create(entreprise=entreprise, tag=tag, defaults={'ajoute_par': request.user})
    return redirect('admin_entreprise_detail', pk=pk)


@admin_required
def retirer_tag_entreprise(request, pk):
    et = get_object_or_404(EntrepriseTag, pk=pk)
    entreprise_pk = et.entreprise.pk
    et.delete()
    return redirect('admin_entreprise_detail', pk=entreprise_pk)


# ============================================
# GESTION TAGS
# ============================================
@admin_required
def tags_list(request):
    tags = TagEntreprise.objects.all()
    if request.method == 'POST':
        TagEntreprise.objects.create(
            nom=request.POST.get('nom'),
            couleur=request.POST.get('couleur', '#14706b'),
            description=request.POST.get('description', ''),
        )
        messages.success(request, "Tag cree")
        return redirect('admin_tags_list')
    return render(request, 'admin_panel/tags_list.html', {'tags': tags})


@admin_required
def tag_delete(request, pk):
    tag = get_object_or_404(TagEntreprise, pk=pk)
    tag.delete()
    return JsonResponse({'status': 'ok'})


# ============================================
# FINANCES
# ============================================
@admin_required
def finances_dashboard(request):
    """Dashboard financier complet"""
    now = timezone.now()
    debut_mois = now.replace(day=1, hour=0, minute=0, second=0)
    debut_annee = now.replace(month=1, day=1, hour=0, minute=0, second=0)

    revenus = {
        'aujourd_hui': float(Paiement.objects.filter(statut='VALIDE', created_at__date=now.date()).aggregate(s=Sum('montant'))['s'] or 0),
        'mois': float(Paiement.objects.filter(statut='VALIDE', created_at__gte=debut_mois).aggregate(s=Sum('montant'))['s'] or 0),
        'annee': float(Paiement.objects.filter(statut='VALIDE', created_at__gte=debut_annee).aggregate(s=Sum('montant'))['s'] or 0),
        'total': float(Paiement.objects.filter(statut='VALIDE').aggregate(s=Sum('montant'))['s'] or 0),
    }

    paiements_recents = Paiement.objects.filter(statut='VALIDE').order_by('-created_at')[:20]
    paiements_attente = Paiement.objects.filter(statut='EN_ATTENTE')

    # Abonnements expirant
    expirent_bientot = Abonnement.objects.filter(
        statut='ACTIF', date_fin__lte=now + timedelta(days=7)
    ).select_related('entreprise', 'forfait')

    return render(request, 'admin_panel/finances.html', {
        'revenus': revenus,
        'paiements_recents': paiements_recents,
        'paiements_attente': paiements_attente,
        'expirent_bientot': expirent_bientot,
    })


# ============================================
# LOGS D'AUDIT
# ============================================
@admin_required
def logs_list(request):
    q = request.GET.get('q', '')
    action_filter = request.GET.get('action', '')
    logs = LogAudit.objects.all().select_related('user')
    if q:
        logs = logs.filter(Q(description__icontains=q) | Q(user__email__icontains=q))
    if action_filter:
        logs = logs.filter(action=action_filter)

    actions = [a[0] for a in LogAudit.ACTIONS]
    return render(request, 'admin_panel/logs_list.html', {
        'logs': logs[:200], 'q': q, 'actions': actions, 'action_filter': action_filter,
    })


# ============================================
# EXPORTS
# ============================================
@admin_required
def export_entreprises_csv(request):
    """Export CSV des entreprises"""
    import csv
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = 'attachment; filename="entreprises.csv"'
    response.write('\ufeff')

    writer = csv.writer(response, delimiter=';')
    writer.writerow(['Nom', 'Secteur', 'Email', 'Telephone', 'Formule', 'Statut', 'Date inscription'])

    for e in Entreprise.objects.all():
        writer.writerow([
            e.nom, e.secteur, e.email_contact, e.telephone,
            e.formule, e.statut, e.date_inscription.strftime('%d/%m/%Y'),
        ])

    log_action(request, 'EXPORT', f"Export CSV de {Entreprise.objects.count()} entreprises")
    return response


@admin_required
def export_paiements_excel(request):
    """Export Excel des paiements"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from io import BytesIO

    wb = Workbook()
    ws = wb.active
    ws.title = "Paiements"

    headers = ['Reference', 'Entreprise', 'Forfait', 'Montant (FCFA)', 'Methode', 'Transaction', 'Statut', 'Date']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='14706B')

    for p in Paiement.objects.all().select_related('entreprise', 'forfait'):
        ws.append([
            p.reference, p.entreprise.nom, p.forfait.nom,
            float(p.montant), p.get_methode_display(),
            p.numero_transaction or '-', p.get_statut_display(),
            p.created_at.strftime('%d/%m/%Y %H:%M'),
        ])

    for col, w in enumerate([18, 30, 20, 15, 25, 25, 15, 18], start=1):
        ws.column_dimensions[chr(64+col)].width = w

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="paiements.xlsx"'

    log_action(request, 'EXPORT', f"Export Excel paiements")
    return response
