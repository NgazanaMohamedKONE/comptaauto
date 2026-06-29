"""
COMPTAAUTO - Pack 1 Essentiel Admin
- Dashboard admin avec graphiques
- Gestion utilisateurs complete
- Gestion entreprises avancee
- Vue financiere
- Logs d'audit
- Impersonate
Usage: python add_admin_pack1.py
"""
import os, sys, subprocess
from pathlib import Path

BASE = Path(__file__).parent

print("\n" + "="*65)
print("  PACK 1 : ESSENTIEL ADMIN")
print("="*65 + "\n")

if os.name == 'nt':
    PY = str(BASE / "venv" / "Scripts" / "python.exe")
else:
    PY = str(BASE / "venv" / "bin" / "python")

print("[1/3] Creation dossiers...")
for d in ["admin_panel", "admin_panel/migrations", "templates/admin_panel"]:
    (BASE / d).mkdir(parents=True, exist_ok=True)
print("      OK\n")

print("[2/3] Generation fichiers...")
FILES = {}

# ============================================
# APP ADMIN_PANEL
# ============================================
FILES["admin_panel/__init__.py"] = ""
FILES["admin_panel/migrations/__init__.py"] = ""

FILES["admin_panel/apps.py"] = '''from django.apps import AppConfig
class AdminPanelConfig(AppConfig):
    default_auto_field = 'django.db.models.BigAutoField'
    name = 'admin_panel'
'''

FILES["admin_panel/models.py"] = '''from django.db import models
from accounts.models import User, Entreprise


class LogAudit(models.Model):
    """Log de toutes les actions importantes"""
    ACTIONS = [
        ('CONNEXION', 'Connexion'),
        ('DECONNEXION', 'Deconnexion'),
        ('CREATION', 'Creation'),
        ('MODIFICATION', 'Modification'),
        ('SUPPRESSION', 'Suppression'),
        ('VALIDATION', 'Validation'),
        ('REFUS', 'Refus'),
        ('IMPERSONATE', 'Impersonification'),
        ('EXPORT', 'Export'),
        ('AUTRE', 'Autre'),
    ]

    user = models.ForeignKey(User, on_delete=models.SET_NULL, null=True, related_name='logs')
    action = models.CharField(max_length=20, choices=ACTIONS)
    objet_type = models.CharField(max_length=100, blank=True, help_text="Type d'objet (User, Entreprise, etc.)")
    objet_id = models.CharField(max_length=50, blank=True)
    description = models.TextField()
    ip_address = models.GenericIPAddressField(null=True, blank=True)
    user_agent = models.TextField(blank=True)
    created_at = models.DateTimeField(auto_now_add=True)

    class Meta:
        db_table = 'logs_audit'
        ordering = ['-created_at']

    def __str__(self):
        return f"{self.user} - {self.action} - {self.created_at}"


class NoteEntreprise(models.Model):
    """Notes internes sur une entreprise (par les admins)"""
    entreprise = models.ForeignKey(Entreprise, on_delete=models.CASCADE, related_name='notes_admin')
    auteur = models.ForeignKey(User, on_delete=models.SET_NULL, null=True)
    contenu = models.TextField()
    importante = models.BooleanField(default=False)
    created_at = models.DateTimeField(auto_now_add=True)

    class Meta:
        db_table = 'notes_entreprise'
        ordering = ['-created_at']


class TagEntreprise(models.Model):
    """Tags pour categoriser les entreprises (VIP, A surveiller, etc.)"""
    nom = models.CharField(max_length=50, unique=True)
    couleur = models.CharField(max_length=20, default='#14706b')
    description = models.TextField(blank=True)

    class Meta:
        db_table = 'tags_entreprise'

    def __str__(self):
        return self.nom


class EntrepriseTag(models.Model):
    """Association entreprise <-> tag"""
    entreprise = models.ForeignKey(Entreprise, on_delete=models.CASCADE, related_name='tags')
    tag = models.ForeignKey(TagEntreprise, on_delete=models.CASCADE)
    ajoute_par = models.ForeignKey(User, on_delete=models.SET_NULL, null=True)
    created_at = models.DateTimeField(auto_now_add=True)

    class Meta:
        db_table = 'entreprise_tags'
        unique_together = ('entreprise', 'tag')


class SessionImpersonate(models.Model):
    """Trace les impersonifications"""
    admin = models.ForeignKey(User, on_delete=models.CASCADE, related_name='impersonations_faites')
    user_cible = models.ForeignKey(User, on_delete=models.CASCADE, related_name='impersonations_subies')
    raison = models.TextField()
    started_at = models.DateTimeField(auto_now_add=True)
    ended_at = models.DateTimeField(null=True, blank=True)

    class Meta:
        db_table = 'sessions_impersonate'
'''

FILES["admin_panel/utils.py"] = '''"""Utilitaires pour l'admin"""
from .models import LogAudit


def log_action(request, action, description, objet=None):
    """Cree un log d'audit"""
    try:
        LogAudit.objects.create(
            user=request.user if request.user.is_authenticated else None,
            action=action,
            objet_type=objet.__class__.__name__ if objet else '',
            objet_id=str(objet.pk) if objet else '',
            description=description,
            ip_address=get_client_ip(request),
            user_agent=request.META.get('HTTP_USER_AGENT', '')[:500],
        )
    except Exception as e:
        print(f"Erreur log: {e}")


def get_client_ip(request):
    """Recupere l'IP du client"""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        return x_forwarded_for.split(',')[0]
    return request.META.get('REMOTE_ADDR')
'''

FILES["admin_panel/views.py"] = '''from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.db.models import Count, Sum, Q, Avg
from django.utils import timezone
from datetime import timedelta, datetime
from django.contrib.auth import login as auth_login, get_user_model
from accounts.models import User, Entreprise
from comptabilite.models import Ecriture
from ocr_app.models import FactureOCR
from communication.models import Ticket, Conversation, DemandeRappel
from abonnements.models import Abonnement, Paiement, Forfait
from .models import LogAudit, NoteEntreprise, TagEntreprise, EntrepriseTag, SessionImpersonate
from .utils import log_action

User = get_user_model()


def admin_required(view_func):
    """Decorator : verifie que c'est un super admin"""
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('login')
        if not request.user.is_super_admin:
            return redirect('dashboard')
        return view_func(request, *args, **kwargs)
    return wrapper


# ============================================
# DASHBOARD ADMIN
# ============================================
@admin_required
def dashboard_admin(request):
    """Dashboard admin enrichi"""
    now = timezone.now()
    debut_mois = now.replace(day=1, hour=0, minute=0, second=0)
    debut_annee = now.replace(month=1, day=1, hour=0, minute=0, second=0)

    # Stats globales
    total_entreprises = Entreprise.objects.count()
    total_users = User.objects.filter(role='ENTREPRISE').count()
    entreprises_actives = Entreprise.objects.filter(statut='ACTIF').count()
    abonnements_actifs = Abonnement.objects.filter(statut='ACTIF').count()

    # Stats financiers
    revenus_total = Paiement.objects.filter(statut='VALIDE').aggregate(s=Sum('montant'))['s'] or 0
    revenus_mois = Paiement.objects.filter(statut='VALIDE', created_at__gte=debut_mois).aggregate(s=Sum('montant'))['s'] or 0
    revenus_annee = Paiement.objects.filter(statut='VALIDE', created_at__gte=debut_annee).aggregate(s=Sum('montant'))['s'] or 0

    # MRR (Monthly Recurring Revenue)
    mrr = 0
    for abo in Abonnement.objects.filter(statut='ACTIF'):
        if abo.forfait.duree_jours <= 31:
            mrr += float(abo.forfait.prix)
        else:
            mrr += float(abo.forfait.prix) * 30 / abo.forfait.duree_jours

    # Alertes
    paiements_en_attente = Paiement.objects.filter(statut='EN_ATTENTE').count()
    tickets_ouverts = Ticket.objects.filter(statut__in=['OUVERT', 'EN_COURS']).count()
    rappels_attente = DemandeRappel.objects.filter(statut='EN_ATTENTE').count()
    abonnements_expirent_bientot = Abonnement.objects.filter(
        statut='ACTIF', date_fin__lte=now + timedelta(days=7)
    ).count()

    # Top entreprises (par nb ecritures)
    top_entreprises = Entreprise.objects.annotate(
        nb_ecr=Count('ecritures')
    ).order_by('-nb_ecr')[:5]

    # Nouvelles entreprises ce mois
    nouvelles_entreprises = Entreprise.objects.filter(date_inscription__gte=debut_mois).count()

    # Repartition par forfait (donut)
    repartition = []
    for f in Forfait.objects.filter(actif=True):
        count = Abonnement.objects.filter(forfait=f, statut='ACTIF').count()
        if count > 0:
            repartition.append({'nom': f.nom, 'count': count, 'couleur': f.couleur})

    # Activite recente
    logs_recents = LogAudit.objects.all()[:10]

    return render(request, 'admin_panel/dashboard.html', {
        'total_entreprises': total_entreprises,
        'total_users': total_users,
        'entreprises_actives': entreprises_actives,
        'abonnements_actifs': abonnements_actifs,
        'revenus_total': revenus_total,
        'revenus_mois': revenus_mois,
        'revenus_annee': revenus_annee,
        'mrr': int(mrr),
        'arr': int(mrr * 12),
        'paiements_en_attente': paiements_en_attente,
        'tickets_ouverts': tickets_ouverts,
        'rappels_attente': rappels_attente,
        'abonnements_expirent_bientot': abonnements_expirent_bientot,
        'top_entreprises': top_entreprises,
        'nouvelles_entreprises': nouvelles_entreprises,
        'repartition': repartition,
        'logs_recents': logs_recents,
    })


@admin_required
def stats_data(request):
    """API : donnees pour graphiques"""
    now = timezone.now()

    # Revenus 12 derniers mois
    mois_labels = []
    mois_revenus = []
    for i in range(11, -1, -1):
        debut = (now.replace(day=1) - timedelta(days=i*30)).replace(day=1)
        fin = (debut + timedelta(days=32)).replace(day=1)
        revenus = Paiement.objects.filter(
            statut='VALIDE', created_at__gte=debut, created_at__lt=fin
        ).aggregate(s=Sum('montant'))['s'] or 0
        mois_labels.append(debut.strftime('%b %Y'))
        mois_revenus.append(float(revenus))

    # Nouveaux clients 12 derniers mois
    mois_clients = []
    for i in range(11, -1, -1):
        debut = (now.replace(day=1) - timedelta(days=i*30)).replace(day=1)
        fin = (debut + timedelta(days=32)).replace(day=1)
        count = Entreprise.objects.filter(date_inscription__gte=debut, date_inscription__lt=fin).count()
        mois_clients.append(count)

    return JsonResponse({
        'revenus': {'labels': mois_labels, 'data': mois_revenus},
        'clients': {'labels': mois_labels, 'data': mois_clients},
    })


# ============================================
# GESTION UTILISATEURS
# ============================================
@admin_required
def users_list(request):
    """Liste de tous les utilisateurs"""
    q = request.GET.get('q', '')
    users = User.objects.all().order_by('-date_joined')
    if q:
        users = users.filter(Q(email__icontains=q) | Q(username__icontains=q) | Q(first_name__icontains=q) | Q(last_name__icontains=q))
    return render(request, 'admin_panel/users_list.html', {'users': users, 'q': q})


@admin_required
def user_detail(request, pk):
    """Detail d'un utilisateur"""
    user = get_object_or_404(User, pk=pk)
    logs = LogAudit.objects.filter(user=user)[:20]
    entreprise = getattr(user, 'entreprise', None)
    return render(request, 'admin_panel/user_detail.html', {
        'user_obj': user, 'logs': logs, 'entreprise': entreprise,
    })


@admin_required
def user_create(request):
    """Creer un utilisateur"""
    if request.method == 'POST':
        try:
            user = User.objects.create_user(
                email=request.POST.get('email'),
                username=request.POST.get('username'),
                password=request.POST.get('password'),
                first_name=request.POST.get('first_name', ''),
                last_name=request.POST.get('last_name', ''),
                telephone=request.POST.get('telephone', ''),
                role=request.POST.get('role', 'ENTREPRISE'),
            )
            if request.POST.get('role') == 'SUPER_ADMIN':
                user.is_staff = True
                user.is_superuser = True
                user.save()
            log_action(request, 'CREATION', f"Creation utilisateur {user.email}", user)
            messages.success(request, f"Utilisateur {user.email} cree")
            return redirect('admin_users_list')
        except Exception as e:
            messages.error(request, f"Erreur: {e}")
    return render(request, 'admin_panel/user_form.html', {'user_obj': None})


@admin_required
def user_edit(request, pk):
    """Modifier un utilisateur"""
    user = get_object_or_404(User, pk=pk)
    if request.method == 'POST':
        user.email = request.POST.get('email', user.email)
        user.username = request.POST.get('username', user.username)
        user.first_name = request.POST.get('first_name', user.first_name)
        user.last_name = request.POST.get('last_name', user.last_name)
        user.telephone = request.POST.get('telephone', user.telephone)
        user.role = request.POST.get('role', user.role)
        user.is_active = request.POST.get('is_active') == 'on'
        user.save()
        log_action(request, 'MODIFICATION', f"Modification utilisateur {user.email}", user)
        messages.success(request, "Utilisateur modifie")
        return redirect('admin_user_detail', pk=user.pk)
    return render(request, 'admin_panel/user_form.html', {'user_obj': user})


@admin_required
def user_delete(request, pk):
    """Supprimer un utilisateur"""
    user = get_object_or_404(User, pk=pk)
    if user == request.user:
        return JsonResponse({'error': 'Vous ne pouvez pas vous supprimer'}, status=400)
    email = user.email
    user.delete()
    log_action(request, 'SUPPRESSION', f"Suppression utilisateur {email}")
    return JsonResponse({'status': 'ok'})


@admin_required
def user_reset_password(request, pk):
    """Reinitialiser mot de passe"""
    user = get_object_or_404(User, pk=pk)
    new_password = request.POST.get('password', 'Temp2026!')
    user.set_password(new_password)
    user.save()
    log_action(request, 'MODIFICATION', f"Reset password {user.email}", user)

    # Notif
    from communication.models import Notification
    Notification.objects.create(
        user=user, type_notif='SYSTEME',
        titre='Mot de passe reinitialise',
        message=f'Votre nouveau mot de passe : {new_password}. Changez-le rapidement !',
    )
    return JsonResponse({'status': 'ok', 'password': new_password})


@admin_required
def user_toggle_active(request, pk):
    """Activer/desactiver un utilisateur"""
    user = get_object_or_404(User, pk=pk)
    user.is_active = not user.is_active
    user.save()
    log_action(request, 'MODIFICATION', f"{'Active' if user.is_active else 'Desactive'} {user.email}", user)
    return JsonResponse({'status': 'ok', 'is_active': user.is_active})


@admin_required
def user_impersonate(request, pk):
    """Se connecter comme un autre utilisateur"""
    user = get_object_or_404(User, pk=pk)
    if user.is_super_admin:
        messages.error(request, "Impossible d'impersonifier un autre super admin")
        return redirect('admin_user_detail', pk=user.pk)

    SessionImpersonate.objects.create(
        admin=request.user, user_cible=user,
        raison=request.POST.get('raison', 'Debug / Support'),
    )
    log_action(request, 'IMPERSONATE', f"Impersonate de {user.email}", user)

    # Sauvegarder l'admin original en session
    request.session['admin_original_id'] = request.user.id
    auth_login(request, user, backend='django.contrib.auth.backends.ModelBackend')
    messages.warning(request, f"Vous etes connecte comme {user.email}. Cliquez sur 'Revenir admin' pour quitter.")
    return redirect('dashboard')


def stop_impersonate(request):
    """Revenir au compte admin original"""
    admin_id = request.session.get('admin_original_id')
    if admin_id:
        admin = User.objects.get(pk=admin_id)
        auth_login(request, admin, backend='django.contrib.auth.backends.ModelBackend')
        del request.session['admin_original_id']
        messages.success(request, "Retour au compte admin")
    return redirect('admin_dashboard')


# ============================================
# GESTION ENTREPRISES AVANCEE
# ============================================
@admin_required
def entreprises_list(request):
    """Liste enrichie des entreprises"""
    q = request.GET.get('q', '')
    statut_filter = request.GET.get('statut', '')
    forfait_filter = request.GET.get('forfait', '')

    entreprises = Entreprise.objects.all().select_related('responsable').prefetch_related('tags__tag', 'abonnements')

    if q:
        entreprises = entreprises.filter(Q(nom__icontains=q) | Q(email_contact__icontains=q) | Q(secteur__icontains=q))
    if statut_filter:
        entreprises = entreprises.filter(statut=statut_filter)

    tags = TagEntreprise.objects.all()

    return render(request, 'admin_panel/entreprises_list.html', {
        'entreprises': entreprises, 'q': q, 'tags': tags,
        'statut_filter': statut_filter,
    })


@admin_required
def entreprise_detail(request, pk):
    """Detail complet d'une entreprise"""
    entreprise = get_object_or_404(Entreprise, pk=pk)

    abonnement_actuel = Abonnement.objects.filter(entreprise=entreprise, statut='ACTIF').first()
    historique_abos = Abonnement.objects.filter(entreprise=entreprise)
    paiements = Paiement.objects.filter(entreprise=entreprise)
    nb_ecritures = Ecriture.objects.filter(entreprise=entreprise).count()
    nb_ocr = FactureOCR.objects.filter(entreprise=entreprise).count()
    tickets = Ticket.objects.filter(entreprise=entreprise)[:10]
    notes = NoteEntreprise.objects.filter(entreprise=entreprise)
    tags = EntrepriseTag.objects.filter(entreprise=entreprise).select_related('tag')
    tags_disponibles = TagEntreprise.objects.exclude(
        id__in=tags.values_list('tag_id', flat=True)
    )

    revenus_entreprise = paiements.filter(statut='VALIDE').aggregate(s=Sum('montant'))['s'] or 0

    return render(request, 'admin_panel/entreprise_detail.html', {
        'entreprise': entreprise,
        'abonnement_actuel': abonnement_actuel,
        'historique_abos': historique_abos,
        'paiements': paiements,
        'nb_ecritures': nb_ecritures,
        'nb_ocr': nb_ocr,
        'tickets': tickets,
        'notes': notes,
        'tags': tags,
        'tags_disponibles': tags_disponibles,
        'revenus_entreprise': revenus_entreprise,
    })


@admin_required
def entreprise_edit(request, pk):
    entreprise = get_object_or_404(Entreprise, pk=pk)
    if request.method == 'POST':
        entreprise.nom = request.POST.get('nom', entreprise.nom)
        entreprise.secteur = request.POST.get('secteur', entreprise.secteur)
        entreprise.adresse = request.POST.get('adresse', entreprise.adresse)
        entreprise.telephone = request.POST.get('telephone', entreprise.telephone)
        entreprise.email_contact = request.POST.get('email_contact', entreprise.email_contact)
        entreprise.formule = request.POST.get('formule', entreprise.formule)
        entreprise.statut = request.POST.get('statut', entreprise.statut)
        entreprise.save()
        log_action(request, 'MODIFICATION', f"Modification entreprise {entreprise.nom}", entreprise)
        messages.success(request, "Entreprise modifiee")
        return redirect('admin_entreprise_detail', pk=entreprise.pk)
    return render(request, 'admin_panel/entreprise_form.html', {'entreprise': entreprise})


@admin_required
def ajouter_note(request, pk):
    entreprise = get_object_or_404(Entreprise, pk=pk)
    if request.method == 'POST':
        NoteEntreprise.objects.create(
            entreprise=entreprise, auteur=request.user,
            contenu=request.POST.get('contenu'),
            importante=request.POST.get('importante') == 'on',
        )
        messages.success(request, "Note ajoutee")
    return redirect('admin_entreprise_detail', pk=pk)


@admin_required
def supprimer_note(request, pk):
    note = get_object_or_404(NoteEntreprise, pk=pk)
    entreprise_pk = note.entreprise.pk
    note.delete()
    return redirect('admin_entreprise_detail', pk=entreprise_pk)


@admin_required
def ajouter_tag_entreprise(request, pk):
    entreprise = get_object_or_404(Entreprise, pk=pk)
    tag_id = request.POST.get('tag_id')
    if tag_id:
        tag = get_object_or_404(TagEntreprise, pk=tag_id)
        EntrepriseTag.objects.get_or_create(entreprise=entreprise, tag=tag, defaults={'ajoute_par': request.user})
    return redirect('admin_entreprise_detail', pk=pk)


@admin_required
def retirer_tag_entreprise(request, pk):
    et = get_object_or_404(EntrepriseTag, pk=pk)
    entreprise_pk = et.entreprise.pk
    et.delete()
    return redirect('admin_entreprise_detail', pk=entreprise_pk)


# ============================================
# GESTION TAGS
# ============================================
@admin_required
def tags_list(request):
    tags = TagEntreprise.objects.all()
    if request.method == 'POST':
        TagEntreprise.objects.create(
            nom=request.POST.get('nom'),
            couleur=request.POST.get('couleur', '#14706b'),
            description=request.POST.get('description', ''),
        )
        messages.success(request, "Tag cree")
        return redirect('admin_tags_list')
    return render(request, 'admin_panel/tags_list.html', {'tags': tags})


@admin_required
def tag_delete(request, pk):
    tag = get_object_or_404(TagEntreprise, pk=pk)
    tag.delete()
    return JsonResponse({'status': 'ok'})


# ============================================
# FINANCES
# ============================================
@admin_required
def finances_dashboard(request):
    """Dashboard financier complet"""
    now = timezone.now()
    debut_mois = now.replace(day=1, hour=0, minute=0, second=0)
    debut_annee = now.replace(month=1, day=1, hour=0, minute=0, second=0)

    revenus = {
        'aujourd_hui': float(Paiement.objects.filter(statut='VALIDE', created_at__date=now.date()).aggregate(s=Sum('montant'))['s'] or 0),
        'mois': float(Paiement.objects.filter(statut='VALIDE', created_at__gte=debut_mois).aggregate(s=Sum('montant'))['s'] or 0),
        'annee': float(Paiement.objects.filter(statut='VALIDE', created_at__gte=debut_annee).aggregate(s=Sum('montant'))['s'] or 0),
        'total': float(Paiement.objects.filter(statut='VALIDE').aggregate(s=Sum('montant'))['s'] or 0),
    }

    paiements_recents = Paiement.objects.filter(statut='VALIDE').order_by('-created_at')[:20]
    paiements_attente = Paiement.objects.filter(statut='EN_ATTENTE')

    # Abonnements expirant
    expirent_bientot = Abonnement.objects.filter(
        statut='ACTIF', date_fin__lte=now + timedelta(days=7)
    ).select_related('entreprise', 'forfait')

    return render(request, 'admin_panel/finances.html', {
        'revenus': revenus,
        'paiements_recents': paiements_recents,
        'paiements_attente': paiements_attente,
        'expirent_bientot': expirent_bientot,
    })


# ============================================
# LOGS D'AUDIT
# ============================================
@admin_required
def logs_list(request):
    q = request.GET.get('q', '')
    action_filter = request.GET.get('action', '')
    logs = LogAudit.objects.all().select_related('user')
    if q:
        logs = logs.filter(Q(description__icontains=q) | Q(user__email__icontains=q))
    if action_filter:
        logs = logs.filter(action=action_filter)

    actions = [a[0] for a in LogAudit.ACTIONS]
    return render(request, 'admin_panel/logs_list.html', {
        'logs': logs[:200], 'q': q, 'actions': actions, 'action_filter': action_filter,
    })


# ============================================
# EXPORTS
# ============================================
@admin_required
def export_entreprises_csv(request):
    """Export CSV des entreprises"""
    import csv
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = 'attachment; filename="entreprises.csv"'
    response.write('\\ufeff')

    writer = csv.writer(response, delimiter=';')
    writer.writerow(['Nom', 'Secteur', 'Email', 'Telephone', 'Formule', 'Statut', 'Date inscription'])

    for e in Entreprise.objects.all():
        writer.writerow([
            e.nom, e.secteur, e.email_contact, e.telephone,
            e.formule, e.statut, e.date_inscription.strftime('%d/%m/%Y'),
        ])

    log_action(request, 'EXPORT', f"Export CSV de {Entreprise.objects.count()} entreprises")
    return response


@admin_required
def export_paiements_excel(request):
    """Export Excel des paiements"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from io import BytesIO

    wb = Workbook()
    ws = wb.active
    ws.title = "Paiements"

    headers = ['Reference', 'Entreprise', 'Forfait', 'Montant (FCFA)', 'Methode', 'Transaction', 'Statut', 'Date']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='14706B')

    for p in Paiement.objects.all().select_related('entreprise', 'forfait'):
        ws.append([
            p.reference, p.entreprise.nom, p.forfait.nom,
            float(p.montant), p.get_methode_display(),
            p.numero_transaction or '-', p.get_statut_display(),
            p.created_at.strftime('%d/%m/%Y %H:%M'),
        ])

    for col, w in enumerate([18, 30, 20, 15, 25, 25, 15, 18], start=1):
        ws.column_dimensions[chr(64+col)].width = w

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="paiements.xlsx"'

    log_action(request, 'EXPORT', f"Export Excel paiements")
    return response
'''

FILES["admin_panel/urls.py"] = '''from django.urls import path
from . import views

urlpatterns = [
    # Dashboard
    path('', views.dashboard_admin, name='admin_dashboard'),
    path('api/stats/', views.stats_data, name='admin_stats_data'),

    # Utilisateurs
    path('users/', views.users_list, name='admin_users_list'),
    path('users/create/', views.user_create, name='admin_user_create'),
    path('users/<int:pk>/', views.user_detail, name='admin_user_detail'),
    path('users/<int:pk>/edit/', views.user_edit, name='admin_user_edit'),
    path('users/<int:pk>/delete/', views.user_delete, name='admin_user_delete'),
    path('users/<int:pk>/reset-password/', views.user_reset_password, name='admin_user_reset_password'),
    path('users/<int:pk>/toggle-active/', views.user_toggle_active, name='admin_user_toggle_active'),
    path('users/<int:pk>/impersonate/', views.user_impersonate, name='admin_user_impersonate'),
    path('stop-impersonate/', views.stop_impersonate, name='admin_stop_impersonate'),

    # Entreprises
    path('entreprises/', views.entreprises_list, name='admin_entreprises_list'),
    path('entreprises/<int:pk>/', views.entreprise_detail, name='admin_entreprise_detail'),
    path('entreprises/<int:pk>/edit/', views.entreprise_edit, name='admin_entreprise_edit'),
    path('entreprises/<int:pk>/note/', views.ajouter_note, name='admin_ajouter_note'),
    path('notes/<int:pk>/delete/', views.supprimer_note, name='admin_supprimer_note'),
    path('entreprises/<int:pk>/tag/', views.ajouter_tag_entreprise, name='admin_ajouter_tag_entreprise'),
    path('entreprise-tags/<int:pk>/delete/', views.retirer_tag_entreprise, name='admin_retirer_tag_entreprise'),

    # Tags
    path('tags/', views.tags_list, name='admin_tags_list'),
    path('tags/<int:pk>/delete/', views.tag_delete, name='admin_tag_delete'),

    # Finances
    path('finances/', views.finances_dashboard, name='admin_finances'),

    # Logs
    path('logs/', views.logs_list, name='admin_logs_list'),

    # Exports
    path('export/entreprises-csv/', views.export_entreprises_csv, name='admin_export_entreprises_csv'),
    path('export/paiements-excel/', views.export_paiements_excel, name='admin_export_paiements_excel'),
]
'''

FILES["admin_panel/admin.py"] = '''from django.contrib import admin
from .models import LogAudit, NoteEntreprise, TagEntreprise, EntrepriseTag, SessionImpersonate

@admin.register(LogAudit)
class LogAuditAdmin(admin.ModelAdmin):
    list_display = ('user', 'action', 'description', 'created_at')
    list_filter = ('action',)
    search_fields = ('description', 'user__email')

@admin.register(NoteEntreprise)
class NoteEntrepriseAdmin(admin.ModelAdmin):
    list_display = ('entreprise', 'auteur', 'importante', 'created_at')

@admin.register(TagEntreprise)
class TagEntrepriseAdmin(admin.ModelAdmin):
    list_display = ('nom', 'couleur')
'''

# ============================================
# TEMPLATES ADMIN
# ============================================

FILES["templates/admin_panel/base_admin.html"] = '''<!DOCTYPE html>
<html lang="fr">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>{% block title %}Admin - ComptaAuto{% endblock %}</title>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&family=Playfair+Display:wght@700&display=swap" rel="stylesheet">
<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/css/bootstrap.min.css" rel="stylesheet">
<link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.3/font/bootstrap-icons.css">
{% load static %}
<link rel="stylesheet" href="{% static 'css/main.css' %}">
<style>
.admin-sidebar { width: 260px; background: #1a3e3a; color: white; min-height: 100vh; padding: 20px; position: fixed; left: 0; top: 0; overflow-y: auto; }
.admin-sidebar .brand-title { color: white; }
.admin-sidebar .nav-item { color: rgba(255,255,255,0.8); padding: 10px 12px; border-radius: 8px; text-decoration: none; display: flex; align-items: center; gap: 10px; margin-bottom: 4px; }
.admin-sidebar .nav-item:hover { background: rgba(255,255,255,0.1); color: white; }
.admin-sidebar .nav-item.active { background: #14706b; color: white; }
.admin-sidebar .sidebar-section { font-size: 0.7rem; color: rgba(255,255,255,0.5); letter-spacing: 1.5px; margin: 20px 0 10px; }
.admin-main { margin-left: 260px; padding: 0; min-height: 100vh; background: #f5f7f7; }
.admin-topbar { background: white; padding: 15px 30px; border-bottom: 1px solid #eee; display: flex; justify-content: space-between; align-items: center; }
.admin-content { padding: 30px; }
.metric-card { background: white; padding: 24px; border-radius: 12px; box-shadow: 0 2px 8px rgba(0,0,0,0.05); border-left: 4px solid #14706b; }
.metric-value { font-size: 2rem; font-weight: 700; font-family: 'Playfair Display', serif; }
.metric-label { color: #6b8884; font-size: 0.9rem; }
</style>
</head>
<body>
<aside class="admin-sidebar">
<div class="logo-wrapper mb-4">
<div class="logo-icon" style="background:#f4b860;">C</div>
<div><h1 class="brand-title" style="color:white;font-size:1.4rem;">ComptaAuto</h1><p class="brand-subtitle" style="color:rgba(255,255,255,0.6);">CONSOLE ADMIN</p></div>
</div>

<div class="sidebar-section">PILOTAGE</div>
<nav>
<a href="{% url 'admin_dashboard' %}" class="nav-item {% if request.resolver_match.url_name == 'admin_dashboard' %}active{% endif %}"><i class="bi bi-grid-1x2"></i> Dashboard</a>
<a href="{% url 'admin_finances' %}" class="nav-item {% if 'finances' in request.path %}active{% endif %}"><i class="bi bi-currency-exchange"></i> Finances</a>
</nav>

<div class="sidebar-section">GESTION</div>
<nav>
<a href="{% url 'admin_entreprises_list' %}" class="nav-item {% if 'entreprise' in request.path %}active{% endif %}"><i class="bi bi-building"></i> Entreprises</a>
<a href="{% url 'admin_users_list' %}" class="nav-item {% if 'users' in request.path %}active{% endif %}"><i class="bi bi-people"></i> Utilisateurs</a>
<a href="{% url 'admin_tags_list' %}" class="nav-item {% if 'tags' in request.path %}active{% endif %}"><i class="bi bi-tags"></i> Tags</a>
</nav>

<div class="sidebar-section">FACTURATION</div>
<nav>
<a href="{% url 'admin_paiements' %}" class="nav-item {% if 'paiements' in request.path %}active{% endif %}"><i class="bi bi-credit-card"></i> Paiements</a>
<a href="{% url 'admin_abonnements' %}" class="nav-item {% if 'admin/abonnements' in request.path %}active{% endif %}"><i class="bi bi-star"></i> Abonnements</a>
</nav>

<div class="sidebar-section">SUPPORT</div>
<nav>
<a href="{% url 'tickets_list' %}" class="nav-item"><i class="bi bi-ticket-perforated"></i> Tickets</a>
<a href="{% url 'messagerie' %}" class="nav-item"><i class="bi bi-chat-dots"></i> Messages</a>
<a href="{% url 'rappels_list' %}" class="nav-item"><i class="bi bi-telephone"></i> Rappels</a>
<a href="{% url 'annonces_list' %}" class="nav-item"><i class="bi bi-megaphone"></i> Annonces</a>
</nav>

<div class="sidebar-section">SYSTEME</div>
<nav>
<a href="{% url 'admin_logs_list' %}" class="nav-item {% if 'logs' in request.path %}active{% endif %}"><i class="bi bi-list-check"></i> Logs audit</a>
<a href="/django-admin/" class="nav-item" target="_blank"><i class="bi bi-gear"></i> Django Admin</a>
</nav>
</aside>

<main class="admin-main">
<header class="admin-topbar">
<div>
<h4 style="margin:0;font-family:'Playfair Display',serif;">{% block page_title %}Dashboard{% endblock %}</h4>
<small class="text-muted">{% block page_subtitle %}{% endblock %}</small>
</div>
<div class="d-flex align-items-center gap-3">
{% if request.session.admin_original_id %}
<a href="{% url 'admin_stop_impersonate' %}" class="btn btn-warning btn-sm"><i class="bi bi-arrow-left"></i> Revenir admin</a>
{% endif %}
<a href="{% url 'notifications_list' %}" style="color:#14706b;text-decoration:none;font-size:1.3rem;"><i class="bi bi-bell-fill"></i></a>
<div class="user-info"><div class="user-avatar" style="background:#1a3e3a;">{{ user.first_name.0|default:user.email.0|upper }}</div><div><strong>{{ user.get_full_name|default:user.email }}</strong><small>Super Admin</small></div></div>
<a href="{% url 'logout' %}" class="btn btn-outline-danger btn-sm">Deconnexion</a>
</div>
</header>

<div class="admin-content">
{% if messages %}{% for m in messages %}<div class="alert alert-{{ m.tags }} alert-dismissible fade show">{{ m }}<button type="button" class="btn-close" data-bs-dismiss="alert"></button></div>{% endfor %}{% endif %}
{% block content %}{% endblock %}
</div>
</main>

<script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/js/bootstrap.bundle.min.js"></script>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
{% block extra_js %}{% endblock %}
</body>
</html>
'''

FILES["templates/admin_panel/dashboard.html"] = '''{% extends 'admin_panel/base_admin.html' %}{% block page_title %}Dashboard Administrateur{% endblock %}{% block page_subtitle %}Vue d'ensemble de la plateforme{% endblock %}{% block content %}

<div class="row g-3 mb-4">
<div class="col-md-3"><div class="metric-card"><div class="metric-label">REVENUS DU MOIS</div><div class="metric-value text-success">{{ revenus_mois|floatformat:0 }} F</div><small class="text-muted">Total : {{ revenus_total|floatformat:0 }} F</small></div></div>
<div class="col-md-3"><div class="metric-card" style="border-left-color:#1a8b85;"><div class="metric-label">MRR (Revenu Mensuel Recurrent)</div><div class="metric-value">{{ mrr|floatformat:0 }} F</div><small class="text-muted">ARR : {{ arr|floatformat:0 }} F</small></div></div>
<div class="col-md-3"><div class="metric-card" style="border-left-color:#f4b860;"><div class="metric-label">ENTREPRISES</div><div class="metric-value">{{ total_entreprises }}</div><small class="text-success">+{{ nouvelles_entreprises }} ce mois</small></div></div>
<div class="col-md-3"><div class="metric-card" style="border-left-color:#9b59b6;"><div class="metric-label">ABONNEMENTS ACTIFS</div><div class="metric-value">{{ abonnements_actifs }}</div><small class="text-muted">/ {{ total_entreprises }} entreprises</small></div></div>
</div>

{% if paiements_en_attente or tickets_ouverts or rappels_attente or abonnements_expirent_bientot %}
<div class="row g-3 mb-4">
{% if paiements_en_attente %}
<div class="col-md-3"><a href="{% url 'admin_paiements' %}" class="text-decoration-none"><div class="metric-card" style="border-left-color:#ffc107;background:#fff8e1;"><div class="metric-label text-warning"><i class="bi bi-exclamation-triangle"></i> A VALIDER</div><div class="metric-value">{{ paiements_en_attente }}</div><small>Paiement(s) en attente</small></div></a></div>
{% endif %}
{% if tickets_ouverts %}
<div class="col-md-3"><a href="{% url 'tickets_list' %}" class="text-decoration-none"><div class="metric-card" style="border-left-color:#dc3545;background:#fce4e4;"><div class="metric-label text-danger"><i class="bi bi-ticket-perforated"></i> TICKETS</div><div class="metric-value">{{ tickets_ouverts }}</div><small>Ouvert(s) / En cours</small></div></a></div>
{% endif %}
{% if rappels_attente %}
<div class="col-md-3"><a href="{% url 'rappels_list' %}" class="text-decoration-none"><div class="metric-card" style="border-left-color:#0dcaf0;background:#e0f7fa;"><div class="metric-label text-info"><i class="bi bi-telephone"></i> RAPPELS</div><div class="metric-value">{{ rappels_attente }}</div><small>A effectuer</small></div></a></div>
{% endif %}
{% if abonnements_expirent_bientot %}
<div class="col-md-3"><a href="{% url 'admin_finances' %}" class="text-decoration-none"><div class="metric-card" style="border-left-color:#fd7e14;background:#fff3e0;"><div class="metric-label text-warning"><i class="bi bi-hourglass"></i> EXPIRATIONS</div><div class="metric-value">{{ abonnements_expirent_bientot }}</div><small>Sous 7 jours</small></div></a></div>
{% endif %}
</div>
{% endif %}

<div class="row g-3 mb-4">
<div class="col-md-8"><div class="card"><div class="card-body">
<h5>Revenus & Nouveaux clients (12 derniers mois)</h5>
<canvas id="chartRevenus" height="100"></canvas>
</div></div></div>

<div class="col-md-4"><div class="card"><div class="card-body">
<h5>Repartition par forfait</h5>
<canvas id="chartForfaits" height="200"></canvas>
</div></div></div>
</div>

<div class="row g-3">
<div class="col-md-7"><div class="card"><div class="card-body">
<h5>Top 5 entreprises actives</h5>
<table class="table">
<thead><tr><th>Entreprise</th><th>Secteur</th><th>Ecritures</th><th></th></tr></thead>
<tbody>
{% for e in top_entreprises %}
<tr><td><strong>{{ e.nom }}</strong></td><td>{{ e.secteur }}</td><td><span class="badge bg-primary">{{ e.nb_ecr }}</span></td><td><a href="{% url 'admin_entreprise_detail' e.pk %}" class="btn btn-sm btn-outline-primary"><i class="bi bi-eye"></i></a></td></tr>
{% endfor %}
</tbody>
</table>
</div></div></div>

<div class="col-md-5"><div class="card"><div class="card-body">
<h5>Activite recente</h5>
{% for l in logs_recents %}
<div class="border-bottom py-2"><small class="text-muted">{{ l.created_at|timesince }}</small><br><strong>{{ l.user.email|default:"Systeme" }}</strong> - {{ l.description|truncatewords:8 }}</div>
{% empty %}<p class="text-muted">Aucune activite</p>{% endfor %}
<a href="{% url 'admin_logs_list' %}" class="btn btn-sm btn-link">Voir tous les logs</a>
</div></div></div>
</div>

{% endblock %}
{% block extra_js %}
<script>
fetch('{% url "admin_stats_data" %}').then(r=>r.json()).then(d=>{
    new Chart(document.getElementById('chartRevenus'), {
        type: 'bar',
        data: {
            labels: d.revenus.labels,
            datasets: [
                {label: 'Revenus (FCFA)', data: d.revenus.data, backgroundColor: '#14706b', yAxisID: 'y'},
                {label: 'Nouveaux clients', data: d.clients.data, type: 'line', borderColor: '#f4b860', backgroundColor: '#f4b860', yAxisID: 'y1', tension: 0.3}
            ]
        },
        options: {
            responsive: true,
            scales: {
                y: {position: 'left', title: {display: true, text: 'Revenus'}},
                y1: {position: 'right', title: {display: true, text: 'Clients'}, grid: {drawOnChartArea: false}}
            }
        }
    });
});

const repData = {{ repartition|safe }};
new Chart(document.getElementById('chartForfaits'), {
    type: 'doughnut',
    data: {
        labels: repData.map(r => r.nom),
        datasets: [{data: repData.map(r => r.count), backgroundColor: repData.map(r => r.couleur)}]
    },
    options: {plugins: {legend: {position: 'bottom'}}}
});
</script>
{% endblock %}
'''

FILES["templates/admin_panel/users_list.html"] = '''{% extends 'admin_panel/base_admin.html' %}{% block page_title %}Utilisateurs{% endblock %}{% block content %}
<div class="d-flex justify-content-between mb-3">
<form method="GET" class="d-flex gap-2"><input type="text" name="q" class="form-control" placeholder="Rechercher..." value="{{ q }}" style="width:300px;"><button class="btn btn-outline-primary"><i class="bi bi-search"></i></button></form>
<a href="{% url 'admin_user_create' %}" class="btn btn-primary"><i class="bi bi-plus"></i> Nouvel utilisateur</a>
</div>
<div class="card"><div class="card-body">
<table class="table">
<thead><tr><th>Email</th><th>Nom</th><th>Role</th><th>Inscription</th><th>Derniere connexion</th><th>Statut</th><th>Actions</th></tr></thead>
<tbody>
{% for u in users %}
<tr>
<td><strong>{{ u.email }}</strong></td>
<td>{{ u.get_full_name|default:"-" }}</td>
<td><span class="badge bg-{% if u.role == 'SUPER_ADMIN' %}warning{% else %}primary{% endif %}">{{ u.get_role_display }}</span></td>
<td>{{ u.date_joined|date:"d/m/Y" }}</td>
<td>{{ u.last_login|date:"d/m/Y H:i"|default:"Jamais" }}</td>
<td>{% if u.is_active %}<span class="badge bg-success">Actif</span>{% else %}<span class="badge bg-danger">Bloque</span>{% endif %}</td>
<td>
<a href="{% url 'admin_user_detail' u.pk %}" class="btn btn-sm btn-outline-primary"><i class="bi bi-eye"></i></a>
<a href="{% url 'admin_user_edit' u.pk %}" class="btn btn-sm btn-outline-secondary"><i class="bi bi-pencil"></i></a>
{% if u != user %}<button class="btn btn-sm btn-outline-danger" onclick="supprimer({{ u.pk }})"><i class="bi bi-trash"></i></button>{% endif %}
</td>
</tr>
{% endfor %}
</tbody>
</table>
</div></div>
<script>
function getCookie(n){const v=`; ${document.cookie}`.split(`; ${n}=`);if(v.length===2)return v.pop().split(';').shift();}
async function supprimer(id){if(!confirm('Supprimer ?'))return;const r=await fetch('/admin-panel/users/'+id+'/delete/',{method:'POST',headers:{'X-CSRFToken':getCookie('csrftoken')}});if(r.ok)location.reload();}
</script>
{% endblock %}
'''

FILES["templates/admin_panel/user_detail.html"] = '''{% extends 'admin_panel/base_admin.html' %}{% block page_title %}{{ user_obj.email }}{% endblock %}{% block content %}
<a href="{% url 'admin_users_list' %}" class="btn btn-link"><i class="bi bi-arrow-left"></i> Retour</a>
<div class="row g-3">
<div class="col-md-4"><div class="card"><div class="card-body text-center">
<div class="user-avatar mx-auto mb-3" style="width:80px;height:80px;font-size:2rem;background:#14706b;color:white;border-radius:50%;display:flex;align-items:center;justify-content:center;">{{ user_obj.first_name.0|default:user_obj.email.0|upper }}</div>
<h4>{{ user_obj.get_full_name|default:user_obj.email }}</h4>
<p class="text-muted">{{ user_obj.email }}</p>
<span class="badge bg-{% if user_obj.role == 'SUPER_ADMIN' %}warning{% else %}primary{% endif %}">{{ user_obj.get_role_display }}</span>
{% if user_obj.is_active %}<span class="badge bg-success">Actif</span>{% else %}<span class="badge bg-danger">Bloque</span>{% endif %}
<hr>
<div class="text-start">
<p><strong>Telephone :</strong> {{ user_obj.telephone|default:"-" }}</p>
<p><strong>Inscription :</strong> {{ user_obj.date_joined|date:"d/m/Y H:i" }}</p>
<p><strong>Derniere connexion :</strong> {{ user_obj.last_login|date:"d/m/Y H:i"|default:"Jamais" }}</p>
</div>
<div class="d-grid gap-2 mt-3">
<a href="{% url 'admin_user_edit' user_obj.pk %}" class="btn btn-primary"><i class="bi bi-pencil"></i> Modifier</a>
<button class="btn btn-warning" onclick="resetPwd({{ user_obj.pk }})"><i class="bi bi-key"></i> Reset password</button>
<button class="btn btn-{% if user_obj.is_active %}danger{% else %}success{% endif %}" onclick="toggleActive({{ user_obj.pk }})">{% if user_obj.is_active %}Bloquer{% else %}Debloquer{% endif %}</button>
{% if not user_obj.is_super_admin %}
<form method="POST" action="{% url 'admin_user_impersonate' user_obj.pk %}">{% csrf_token %}<input type="hidden" name="raison" value="Debug admin"><button type="submit" class="btn btn-outline-warning" onclick="return confirm('Se connecter comme {{ user_obj.email }} ?')"><i class="bi bi-person-badge"></i> Se connecter comme...</button></form>
{% endif %}
</div>
</div></div></div>

<div class="col-md-8">
{% if entreprise %}
<div class="card mb-3"><div class="card-body">
<h5><i class="bi bi-building"></i> Entreprise associee</h5>
<p><strong>{{ entreprise.nom }}</strong> - {{ entreprise.secteur }}</p>
<a href="{% url 'admin_entreprise_detail' entreprise.pk %}" class="btn btn-sm btn-outline-primary">Voir details</a>
</div></div>
{% endif %}

<div class="card"><div class="card-body">
<h5>Activite recente</h5>
{% for l in logs %}
<div class="border-bottom py-2"><small class="text-muted">{{ l.created_at|date:"d/m/Y H:i" }}</small> - <span class="badge bg-secondary">{{ l.get_action_display }}</span> {{ l.description }}</div>
{% empty %}<p class="text-muted">Aucune activite</p>{% endfor %}
</div></div>
</div>
</div>
<script>
function getCookie(n){const v=`; ${document.cookie}`.split(`; ${n}=`);if(v.length===2)return v.pop().split(';').shift();}
async function resetPwd(id){const p=prompt('Nouveau mot de passe :','Temp2026!');if(!p)return;const fd=new FormData();fd.append('password',p);const r=await fetch('/admin-panel/users/'+id+'/reset-password/',{method:'POST',headers:{'X-CSRFToken':getCookie('csrftoken')},body:fd});const j=await r.json();if(r.ok)alert('Mot de passe : '+j.password);}
async function toggleActive(id){const r=await fetch('/admin-panel/users/'+id+'/toggle-active/',{method:'POST',headers:{'X-CSRFToken':getCookie('csrftoken')}});if(r.ok)location.reload();}
</script>
{% endblock %}
'''

FILES["templates/admin_panel/user_form.html"] = '''{% extends 'admin_panel/base_admin.html' %}{% block page_title %}{% if user_obj %}Modifier{% else %}Nouvel{% endif %} utilisateur{% endblock %}{% block content %}
<div class="card"><div class="card-body">
<form method="POST">{% csrf_token %}
<div class="row">
<div class="col-md-6 mb-3"><label>Email</label><input type="email" name="email" class="form-control" value="{{ user_obj.email }}" required></div>
<div class="col-md-6 mb-3"><label>Username</label><input type="text" name="username" class="form-control" value="{{ user_obj.username }}" required></div>
<div class="col-md-6 mb-3"><label>Prenom</label><input type="text" name="first_name" class="form-control" value="{{ user_obj.first_name }}"></div>
<div class="col-md-6 mb-3"><label>Nom</label><input type="text" name="last_name" class="form-control" value="{{ user_obj.last_name }}"></div>
<div class="col-md-6 mb-3"><label>Telephone</label><input type="text" name="telephone" class="form-control" value="{{ user_obj.telephone }}"></div>
<div class="col-md-6 mb-3"><label>Role</label><select name="role" class="form-select"><option value="ENTREPRISE" {% if user_obj.role == 'ENTREPRISE' %}selected{% endif %}>Entreprise</option><option value="SUPER_ADMIN" {% if user_obj.role == 'SUPER_ADMIN' %}selected{% endif %}>Super Admin</option></select></div>
{% if not user_obj %}<div class="col-md-6 mb-3"><label>Mot de passe</label><input type="password" name="password" class="form-control" required minlength="8"></div>{% endif %}
{% if user_obj %}<div class="col-12 mb-3"><label><input type="checkbox" name="is_active" {% if user_obj.is_active %}checked{% endif %}> Compte actif</label></div>{% endif %}
</div>
<button type="submit" class="btn btn-primary"><i class="bi bi-check"></i> Enregistrer</button>
<a href="{% url 'admin_users_list' %}" class="btn btn-secondary">Annuler</a>
</form>
</div></div>
{% endblock %}
'''

FILES["templates/admin_panel/entreprises_list.html"] = '''{% extends 'admin_panel/base_admin.html' %}{% block page_title %}Entreprises{% endblock %}{% block content %}
<div class="d-flex justify-content-between mb-3">
<form method="GET" class="d-flex gap-2"><input type="text" name="q" class="form-control" placeholder="Rechercher..." value="{{ q }}" style="width:300px;"><select name="statut" class="form-select" style="width:150px;"><option value="">Tous</option><option value="ACTIF" {% if statut_filter == 'ACTIF' %}selected{% endif %}>Actifs</option><option value="SUSPENDU" {% if statut_filter == 'SUSPENDU' %}selected{% endif %}>Suspendus</option></select><button class="btn btn-outline-primary"><i class="bi bi-search"></i></button></form>
<a href="{% url 'admin_export_entreprises_csv' %}" class="btn btn-outline-success"><i class="bi bi-file-earmark-spreadsheet"></i> Export CSV</a>
</div>
<div class="card"><div class="card-body">
<table class="table">
<thead><tr><th>Entreprise</th><th>Secteur</th><th>Responsable</th><th>Tags</th><th>Formule</th><th>Statut</th><th>Inscription</th><th></th></tr></thead>
<tbody>
{% for e in entreprises %}
<tr>
<td><strong>{{ e.nom }}</strong><br><small class="text-muted">{{ e.email_contact }}</small></td>
<td>{{ e.secteur }}</td>
<td>{{ e.responsable.get_full_name }}</td>
<td>{% for t in e.tags.all %}<span class="badge me-1" style="background:{{ t.tag.couleur }};">{{ t.tag.nom }}</span>{% endfor %}</td>
<td><span class="badge bg-light text-dark">{{ e.get_formule_display }}</span></td>
<td>{% if e.statut == 'ACTIF' %}<span class="badge bg-success">Actif</span>{% else %}<span class="badge bg-danger">{{ e.get_statut_display }}</span>{% endif %}</td>
<td>{{ e.date_inscription|date:"d/m/Y" }}</td>
<td><a href="{% url 'admin_entreprise_detail' e.pk %}" class="btn btn-sm btn-outline-primary"><i class="bi bi-eye"></i></a></td>
</tr>
{% endfor %}
</tbody>
</table>
</div></div>
{% endblock %}
'''

FILES["templates/admin_panel/entreprise_detail.html"] = '''{% extends 'admin_panel/base_admin.html' %}{% block page_title %}{{ entreprise.nom }}{% endblock %}{% block content %}
<a href="{% url 'admin_entreprises_list' %}" class="btn btn-link"><i class="bi bi-arrow-left"></i> Retour</a>

<div class="row g-3 mb-3">
<div class="col-md-3"><div class="metric-card"><div class="metric-label">REVENUS GENERES</div><div class="metric-value text-success">{{ revenus_entreprise|floatformat:0 }} F</div></div></div>
<div class="col-md-3"><div class="metric-card"><div class="metric-label">ECRITURES</div><div class="metric-value">{{ nb_ecritures }}</div></div></div>
<div class="col-md-3"><div class="metric-card"><div class="metric-label">FACTURES OCR</div><div class="metric-value">{{ nb_ocr }}</div></div></div>
<div class="col-md-3"><div class="metric-card"><div class="metric-label">TICKETS</div><div class="metric-value">{{ tickets|length }}</div></div></div>
</div>

<div class="row g-3">
<div class="col-md-4">
<div class="card mb-3"><div class="card-body">
<h5><i class="bi bi-building"></i> Informations</h5>
<p><strong>Secteur :</strong> {{ entreprise.secteur }}</p>
<p><strong>Email :</strong> {{ entreprise.email_contact }}</p>
<p><strong>Telephone :</strong> {{ entreprise.telephone|default:"-" }}</p>
<p><strong>Adresse :</strong> {{ entreprise.adresse|default:"-" }}</p>
<p><strong>Formule :</strong> {{ entreprise.get_formule_display }}</p>
<p><strong>Statut :</strong> <span class="badge bg-{% if entreprise.statut == 'ACTIF' %}success{% else %}danger{% endif %}">{{ entreprise.get_statut_display }}</span></p>
<a href="{% url 'admin_entreprise_edit' entreprise.pk %}" class="btn btn-primary btn-sm"><i class="bi bi-pencil"></i> Modifier</a>
</div></div>

<div class="card mb-3"><div class="card-body">
<h5><i class="bi bi-tags"></i> Tags</h5>
{% for t in tags %}<span class="badge me-1 mb-1" style="background:{{ t.tag.couleur }};">{{ t.tag.nom }} <a href="javascript:retirerTag({{ t.pk }})" style="color:white;text-decoration:none;">x</a></span>{% empty %}<p class="text-muted small">Aucun tag</p>{% endfor %}
{% if tags_disponibles %}
<form method="POST" action="{% url 'admin_ajouter_tag_entreprise' entreprise.pk %}" class="mt-2">{% csrf_token %}
<div class="input-group input-group-sm">
<select name="tag_id" class="form-select form-select-sm"><option value="">Ajouter un tag...</option>{% for t in tags_disponibles %}<option value="{{ t.pk }}">{{ t.nom }}</option>{% endfor %}</select>
<button class="btn btn-sm btn-primary">+</button>
</div>
</form>
{% endif %}
</div></div>

<div class="card"><div class="card-body">
<h5><i class="bi bi-sticky"></i> Notes internes</h5>
<form method="POST" action="{% url 'admin_ajouter_note' entreprise.pk %}" class="mb-3">{% csrf_token %}
<textarea name="contenu" class="form-control mb-2" rows="2" placeholder="Ajouter une note..." required></textarea>
<label><input type="checkbox" name="importante"> Importante</label>
<button class="btn btn-sm btn-primary float-end"><i class="bi bi-plus"></i> Ajouter</button>
</form>
{% for n in notes %}
<div class="border-bottom py-2 {% if n.importante %}bg-warning bg-opacity-10 p-2 rounded{% endif %}">
{% if n.importante %}<span class="badge bg-warning text-dark">Important</span>{% endif %}
<small class="text-muted">{{ n.created_at|date:"d/m H:i" }} - {{ n.auteur.email }}</small>
<a href="{% url 'admin_supprimer_note' n.pk %}" class="float-end text-danger" onclick="return confirm('Supprimer ?')"><i class="bi bi-x"></i></a>
<p>{{ n.contenu }}</p>
</div>
{% empty %}<p class="text-muted small">Aucune note</p>{% endfor %}
</div></div>
</div>

<div class="col-md-8">
{% if abonnement_actuel %}
<div class="card mb-3" style="border-top: 4px solid {{ abonnement_actuel.forfait.couleur }};"><div class="card-body">
<h5><i class="bi bi-star"></i> Abonnement actuel</h5>
<div class="row">
<div class="col-md-6"><p><strong>Forfait :</strong> {{ abonnement_actuel.forfait.nom }}</p><p><strong>Debut :</strong> {{ abonnement_actuel.date_debut|date:"d/m/Y" }}</p></div>
<div class="col-md-6"><p><strong>Fin :</strong> {{ abonnement_actuel.date_fin|date:"d/m/Y" }}</p><p><strong>Restant :</strong> <span class="badge bg-success">{{ abonnement_actuel.jours_restants }} jours</span></p></div>
</div>
</div></div>
{% endif %}

<div class="card mb-3"><div class="card-body">
<h5><i class="bi bi-credit-card"></i> Historique paiements</h5>
<table class="table table-sm">
<thead><tr><th>Ref</th><th>Forfait</th><th>Montant</th><th>Statut</th><th>Date</th></tr></thead>
<tbody>
{% for p in paiements %}<tr><td>{{ p.reference }}</td><td>{{ p.forfait.nom }}</td><td>{{ p.montant|floatformat:0 }} F</td><td><span class="badge bg-{% if p.statut == 'VALIDE' %}success{% else %}warning{% endif %}">{{ p.get_statut_display }}</span></td><td>{{ p.created_at|date:"d/m/Y" }}</td></tr>{% empty %}<tr><td colspan="5" class="text-muted">Aucun</td></tr>{% endfor %}
</tbody>
</table>
</div></div>

<div class="card"><div class="card-body">
<h5><i class="bi bi-ticket-perforated"></i> Tickets recents</h5>
{% for t in tickets %}<div class="border-bottom py-2"><a href="{% url 'ticket_detail' t.pk %}">{{ t.numero }} - {{ t.sujet }}</a> <span class="badge bg-{% if t.statut == 'RESOLU' %}success{% elif t.statut == 'OUVERT' %}warning{% else %}primary{% endif %}">{{ t.get_statut_display }}</span></div>{% empty %}<p class="text-muted small">Aucun ticket</p>{% endfor %}
</div></div>
</div>
</div>
<script>
function getCookie(n){const v=`; ${document.cookie}`.split(`; ${n}=`);if(v.length===2)return v.pop().split(';').shift();}
async function retirerTag(id){await fetch('/admin-panel/entreprise-tags/'+id+'/delete/',{method:'POST',headers:{'X-CSRFToken':getCookie('csrftoken')}});location.reload();}
</script>
{% endblock %}
'''

FILES["templates/admin_panel/entreprise_form.html"] = '''{% extends 'admin_panel/base_admin.html' %}{% block page_title %}Modifier {{ entreprise.nom }}{% endblock %}{% block content %}
<div class="card"><div class="card-body">
<form method="POST">{% csrf_token %}
<div class="row">
<div class="col-md-6 mb-3"><label>Nom</label><input type="text" name="nom" class="form-control" value="{{ entreprise.nom }}" required></div>
<div class="col-md-6 mb-3"><label>Secteur</label><input type="text" name="secteur" class="form-control" value="{{ entreprise.secteur }}"></div>
<div class="col-md-6 mb-3"><label>Email contact</label><input type="email" name="email_contact" class="form-control" value="{{ entreprise.email_contact }}"></div>
<div class="col-md-6 mb-3"><label>Telephone</label><input type="text" name="telephone" class="form-control" value="{{ entreprise.telephone }}"></div>
<div class="col-12 mb-3"><label>Adresse</label><textarea name="adresse" class="form-control" rows="2">{{ entreprise.adresse }}</textarea></div>
<div class="col-md-6 mb-3"><label>Formule</label><select name="formule" class="form-select"><option value="STARTER" {% if entreprise.formule == 'STARTER' %}selected{% endif %}>Starter</option><option value="BUSINESS" {% if entreprise.formule == 'BUSINESS' %}selected{% endif %}>Business</option><option value="ENTERPRISE" {% if entreprise.formule == 'ENTERPRISE' %}selected{% endif %}>Enterprise</option></select></div>
<div class="col-md-6 mb-3"><label>Statut</label><select name="statut" class="form-select"><option value="ACTIF" {% if entreprise.statut == 'ACTIF' %}selected{% endif %}>Actif</option><option value="SUSPENDU" {% if entreprise.statut == 'SUSPENDU' %}selected{% endif %}>Suspendu</option></select></div>
</div>
<button type="submit" class="btn btn-primary"><i class="bi bi-check"></i> Enregistrer</button>
<a href="{% url 'admin_entreprise_detail' entreprise.pk %}" class="btn btn-secondary">Annuler</a>
</form>
</div></div>
{% endblock %}
'''

FILES["templates/admin_panel/tags_list.html"] = '''{% extends 'admin_panel/base_admin.html' %}{% block page_title %}Tags{% endblock %}{% block content %}
<div class="row">
<div class="col-md-4"><div class="card"><div class="card-body">
<h5>Nouveau tag</h5>
<form method="POST">{% csrf_token %}
<div class="mb-2"><label>Nom</label><input type="text" name="nom" class="form-control" required></div>
<div class="mb-2"><label>Couleur</label><input type="color" name="couleur" class="form-control" value="#14706b"></div>
<div class="mb-2"><label>Description</label><textarea name="description" class="form-control" rows="2"></textarea></div>
<button type="submit" class="btn btn-primary"><i class="bi bi-plus"></i> Creer</button>
</form>
</div></div></div>
<div class="col-md-8"><div class="card"><div class="card-body">
<h5>Tags existants</h5>
{% for t in tags %}
<div class="border-bottom py-2 d-flex justify-content-between align-items-center">
<div><span class="badge" style="background:{{ t.couleur }};">{{ t.nom }}</span> <small class="text-muted">{{ t.description }}</small></div>
<button class="btn btn-sm btn-outline-danger" onclick="del({{ t.pk }})"><i class="bi bi-trash"></i></button>
</div>
{% empty %}<p class="text-muted">Aucun tag</p>{% endfor %}
</div></div></div>
</div>
<script>
function getCookie(n){const v=`; ${document.cookie}`.split(`; ${n}=`);if(v.length===2)return v.pop().split(';').shift();}
async function del(id){if(!confirm('Supprimer ?'))return;await fetch('/admin-panel/tags/'+id+'/delete/',{method:'POST',headers:{'X-CSRFToken':getCookie('csrftoken')}});location.reload();}
</script>
{% endblock %}
'''

FILES["templates/admin_panel/finances.html"] = '''{% extends 'admin_panel/base_admin.html' %}{% block page_title %}Finances{% endblock %}{% block content %}
<div class="row g-3 mb-4">
<div class="col-md-3"><div class="metric-card"><div class="metric-label">AUJOURD'HUI</div><div class="metric-value text-success">{{ revenus.aujourd_hui|floatformat:0 }} F</div></div></div>
<div class="col-md-3"><div class="metric-card" style="border-left-color:#1a8b85;"><div class="metric-label">CE MOIS</div><div class="metric-value">{{ revenus.mois|floatformat:0 }} F</div></div></div>
<div class="col-md-3"><div class="metric-card" style="border-left-color:#f4b860;"><div class="metric-label">CETTE ANNEE</div><div class="metric-value">{{ revenus.annee|floatformat:0 }} F</div></div></div>
<div class="col-md-3"><div class="metric-card" style="border-left-color:#9b59b6;"><div class="metric-label">TOTAL</div><div class="metric-value">{{ revenus.total|floatformat:0 }} F</div></div></div>
</div>

<div class="d-flex justify-content-end mb-3"><a href="{% url 'admin_export_paiements_excel' %}" class="btn btn-success"><i class="bi bi-file-earmark-excel"></i> Export Excel</a></div>

<div class="row g-3">
<div class="col-md-7"><div class="card"><div class="card-body">
<h5>Paiements recents (valides)</h5>
<table class="table table-sm">
<thead><tr><th>Ref</th><th>Entreprise</th><th>Forfait</th><th>Montant</th><th>Date</th></tr></thead>
<tbody>
{% for p in paiements_recents %}<tr><td>{{ p.reference }}</td><td>{{ p.entreprise.nom }}</td><td>{{ p.forfait.nom }}</td><td><strong>{{ p.montant|floatformat:0 }} F</strong></td><td>{{ p.created_at|date:"d/m/Y" }}</td></tr>{% endfor %}
</tbody>
</table>
</div></div></div>

<div class="col-md-5">
{% if paiements_attente %}
<div class="card mb-3 border-warning"><div class="card-body">
<h5 class="text-warning"><i class="bi bi-exclamation-triangle"></i> A valider ({{ paiements_attente|length }})</h5>
{% for p in paiements_attente %}<div class="border-bottom py-2"><a href="{% url 'admin_paiements' %}">{{ p.reference }}</a> - {{ p.entreprise.nom }} - <strong>{{ p.montant|floatformat:0 }} F</strong></div>{% endfor %}
</div></div>
{% endif %}

{% if expirent_bientot %}
<div class="card border-danger"><div class="card-body">
<h5 class="text-danger"><i class="bi bi-hourglass"></i> Expirent sous 7 jours</h5>
{% for a in expirent_bientot %}<div class="border-bottom py-2"><strong>{{ a.entreprise.nom }}</strong> - {{ a.forfait.nom }} <span class="badge bg-warning">{{ a.jours_restants }}j</span></div>{% endfor %}
</div></div>
{% endif %}
</div>
</div>
{% endblock %}
'''

FILES["templates/admin_panel/logs_list.html"] = '''{% extends 'admin_panel/base_admin.html' %}{% block page_title %}Logs d'audit{% endblock %}{% block content %}
<form method="GET" class="d-flex gap-2 mb-3">
<input type="text" name="q" class="form-control" placeholder="Rechercher..." value="{{ q }}" style="width:300px;">
<select name="action" class="form-select" style="width:200px;"><option value="">Toutes actions</option>{% for a in actions %}<option value="{{ a }}" {% if action_filter == a %}selected{% endif %}>{{ a }}</option>{% endfor %}</select>
<button class="btn btn-primary"><i class="bi bi-search"></i></button>
</form>
<div class="card"><div class="card-body">
<table class="table table-sm">
<thead><tr><th>Date</th><th>Utilisateur</th><th>Action</th><th>Objet</th><th>Description</th><th>IP</th></tr></thead>
<tbody>
{% for l in logs %}<tr><td><small>{{ l.created_at|date:"d/m/Y H:i:s" }}</small></td><td>{{ l.user.email|default:"Systeme" }}</td><td><span class="badge bg-secondary">{{ l.get_action_display }}</span></td><td><small>{{ l.objet_type }} {% if l.objet_id %}#{{ l.objet_id }}{% endif %}</small></td><td>{{ l.description }}</td><td><small>{{ l.ip_address|default:"-" }}</small></td></tr>{% empty %}<tr><td colspan="6" class="text-center text-muted">Aucun log</td></tr>{% endfor %}
</tbody>
</table>
</div></div>
{% endblock %}
'''

# ============================================
# UPDATE SETTINGS + URLS
# ============================================
FILES["comptaauto/settings.py"] = '''from pathlib import Path
import environ, os

BASE_DIR = Path(__file__).resolve().parent.parent
env = environ.Env(DEBUG=(bool, False))
environ.Env.read_env(os.path.join(BASE_DIR, '.env'))

SECRET_KEY = env('SECRET_KEY', default='dev-key-change-me')
DEBUG = env('DEBUG', default=True)
ALLOWED_HOSTS = env.list('ALLOWED_HOSTS', default=['localhost', '127.0.0.1', '*'])

INSTALLED_APPS = [
    'django.contrib.admin','django.contrib.auth','django.contrib.contenttypes',
    'django.contrib.sessions','django.contrib.messages','django.contrib.staticfiles',
    'accounts','comptabilite','ocr_app','dashboard',
    'rapprochement','reporting','alertes','communication','abonnements',
    'admin_panel',
]

MIDDLEWARE = [
    'django.middleware.security.SecurityMiddleware',
    'whitenoise.middleware.WhiteNoiseMiddleware',
    'django.contrib.sessions.middleware.SessionMiddleware',
    'django.middleware.common.CommonMiddleware',
    'django.middleware.csrf.CsrfViewMiddleware',
    'django.contrib.auth.middleware.AuthenticationMiddleware',
    'django.contrib.messages.middleware.MessageMiddleware',
    'django.middleware.clickjacking.XFrameOptionsMiddleware',
    'abonnements.middleware.AbonnementMiddleware',
]

ROOT_URLCONF = 'comptaauto.urls'
TEMPLATES = [{
    'BACKEND': 'django.template.backends.django.DjangoTemplates',
    'DIRS': [BASE_DIR / 'templates'],
    'APP_DIRS': True,
    'OPTIONS': {'context_processors': [
        'django.template.context_processors.debug',
        'django.template.context_processors.request',
        'django.contrib.auth.context_processors.auth',
        'django.contrib.messages.context_processors.messages',
        'abonnements.middleware.context_abonnement',
    ]},
}]
WSGI_APPLICATION = 'comptaauto.wsgi.application'

DATABASES = {'default': {'ENGINE': 'django.db.backends.sqlite3', 'NAME': BASE_DIR / 'db.sqlite3'}}

AUTH_USER_MODEL = 'accounts.User'
AUTH_PASSWORD_VALIDATORS = [{'NAME': 'django.contrib.auth.password_validation.MinimumLengthValidator'}]

LANGUAGE_CODE = 'fr-fr'
TIME_ZONE = 'Africa/Abidjan'
USE_I18N = True
USE_TZ = True

STATIC_URL = '/static/'
STATIC_ROOT = BASE_DIR / 'staticfiles'
STATICFILES_DIRS = [BASE_DIR / 'static']
STATICFILES_STORAGE = 'whitenoise.storage.CompressedManifestStaticFilesStorage'
MEDIA_URL = '/media/'
MEDIA_ROOT = BASE_DIR / 'media'
DEFAULT_AUTO_FIELD = 'django.db.models.BigAutoField'
LOGIN_URL = '/login/'
LOGIN_REDIRECT_URL = '/dashboard/'
LOGOUT_REDIRECT_URL = '/login/'
OCR_SPACE_API_KEY = env('OCR_SPACE_API_KEY', default='helloworld')
'''

FILES["comptaauto/urls.py"] = '''from django.contrib import admin
from django.urls import path, include
from django.conf import settings
from django.conf.urls.static import static
from django.shortcuts import redirect

urlpatterns = [
    path('django-admin/', admin.site.urls),
    path('', lambda r: redirect('login')),
    path('', include('accounts.urls')),
    path('dashboard/', include('dashboard.urls')),
    path('admin-panel/', include('admin_panel.urls')),
    path('comptabilite/', include('comptabilite.urls')),
    path('ocr/', include('ocr_app.urls')),
    path('rapprochement/', include('rapprochement.urls')),
    path('reporting/', include('reporting.urls')),
    path('alertes/', include('alertes.urls')),
    path('communication/', include('communication.urls')),
    path('abonnements/', include('abonnements.urls')),
]

if settings.DEBUG:
    urlpatterns += static(settings.MEDIA_URL, document_root=settings.MEDIA_ROOT)
    urlpatterns += static(settings.STATIC_URL, document_root=settings.STATIC_ROOT)
'''

# Ecriture
for fp, content in FILES.items():
    p = BASE / fp
    p.parent.mkdir(parents=True, exist_ok=True)
    with open(p, 'w', encoding='utf-8') as f:
        f.write(content)
    print(f"      [OK] {fp}")

print(f"\n      {len(FILES)} fichiers crees/mis a jour\n")

# Migrations
print("[3/3] Migrations base de donnees...")
subprocess.run([PY, "manage.py", "makemigrations", "admin_panel"], check=True)
subprocess.run([PY, "manage.py", "migrate"], check=True)
print("      OK")

print("\n" + "="*65)
print("  PACK 1 ADMIN INSTALLE !")
print("="*65)
print(f"\nRelancer le serveur :")
print(f"   {PY} manage.py runserver")
print(f"\nNouveau dashboard admin : http://127.0.0.1:8000/admin-panel/")
print(f"\nConnectez-vous comme : admin@comptaauto.ci / Admin2026!")
print(f"\nFonctionnalites Pack 1 :")
print(f"  - Dashboard avec graphiques revenus + MRR")
print(f"  - Gestion utilisateurs (CRUD + reset + bloquer + impersonate)")
print(f"  - Gestion entreprises avec notes + tags")
print(f"  - Finances (revenus jour/mois/annee)")
print(f"  - Logs d'audit complets")
print(f"  - Exports CSV/Excel")
print(f"\nUne fois teste, dites 'Pack 1 OK' pour le Pack 2 (Avance)")
print("="*65 + "\n")