"""
COMPTAAUTO - Ajout des modules manquants
- Rapprochement bancaire
- Reporting financier (Bilan + CR + Grand Livre)
- Alertes configurables
Usage: python add_missing_modules.py
"""
import os, sys, subprocess
from pathlib import Path

BASE = Path(__file__).parent

print("\n" + "="*65)
print("  AJOUT MODULES MANQUANTS - Rapprochement + Reporting + Alertes")
print("="*65 + "\n")

# Detection venv
if os.name == 'nt':
    PY = str(BASE / "venv" / "Scripts" / "python.exe")
    PIP = str(BASE / "venv" / "Scripts" / "pip.exe")
else:
    PY = str(BASE / "venv" / "bin" / "python")
    PIP = str(BASE / "venv" / "bin" / "pip")

# ============ INSTALL ============
print("[1/4] Installation dependances supplementaires...")
subprocess.run([PIP, "install", "--quiet", "python-Levenshtein", "reportlab", "openpyxl"], check=True)
print("      OK\n")

# ============ DOSSIERS ============
print("[2/4] Creation dossiers...")
for d in ["rapprochement", "rapprochement/migrations",
          "reporting", "reporting/migrations",
          "alertes", "alertes/migrations",
          "templates/rapprochement", "templates/reporting", "templates/alertes"]:
    (BASE / d).mkdir(parents=True, exist_ok=True)
print("      OK\n")

# ============ FICHIERS ============
print("[3/4] Generation fichiers...")
FILES = {}

# ============================================
# RAPPROCHEMENT BANCAIRE
# ============================================
FILES["rapprochement/__init__.py"] = ""
FILES["rapprochement/migrations/__init__.py"] = ""

FILES["rapprochement/apps.py"] = '''from django.apps import AppConfig
class RapprochementConfig(AppConfig):
    default_auto_field = 'django.db.models.BigAutoField'
    name = 'rapprochement'
'''

FILES["rapprochement/models.py"] = '''from django.db import models
from accounts.models import Entreprise
from comptabilite.models import Ecriture

class ReleveBancaire(models.Model):
    entreprise = models.ForeignKey(Entreprise, on_delete=models.CASCADE, related_name='releves')
    nom = models.CharField(max_length=200)
    banque = models.CharField(max_length=100, default='Banque')
    date_import = models.DateTimeField(auto_now_add=True)
    nb_operations = models.IntegerField(default=0)
    nb_rapprochees = models.IntegerField(default=0)
    class Meta:
        db_table = 'releves_bancaires'
        ordering = ['-date_import']
    def __str__(self):
        return f"{self.banque} - {self.nom}"

class OperationBancaire(models.Model):
    SENS = [('DEBIT', 'Debit'), ('CREDIT', 'Credit')]
    releve = models.ForeignKey(ReleveBancaire, on_delete=models.CASCADE, related_name='operations')
    entreprise = models.ForeignKey(Entreprise, on_delete=models.CASCADE)
    date_operation = models.DateField()
    libelle = models.CharField(max_length=300)
    montant = models.DecimalField(max_digits=15, decimal_places=2)
    sens = models.CharField(max_length=10, choices=SENS, default='DEBIT')
    ecriture_associee = models.ForeignKey(Ecriture, null=True, blank=True, on_delete=models.SET_NULL, related_name='operations_rapprochees')
    score_rapprochement = models.FloatField(default=0.0)
    rapprochee = models.BooleanField(default=False)
    suggeree = models.BooleanField(default=False)
    class Meta:
        db_table = 'operations_bancaires'
        ordering = ['-date_operation']
    def __str__(self):
        return f"{self.date_operation} - {self.libelle} - {self.montant}"
'''

FILES["rapprochement/matcher.py"] = '''"""
Algorithme de rapprochement bancaire
Score = Montant (0.5) + Date +/-3j (0.3) + Libelle Levenshtein (0.2)
"""
from datetime import timedelta
try:
    from Levenshtein import ratio as lev_ratio
except ImportError:
    def lev_ratio(a, b):
        # Fallback simple si Levenshtein pas installe
        a, b = a.lower(), b.lower()
        if not a or not b: return 0.0
        common = sum(1 for c in a if c in b)
        return common / max(len(a), len(b))

def calculate_score(operation, ecriture):
    """Calcule le score de rapprochement entre une operation bancaire et une ecriture"""
    score = 0.0

    # 1. Montant (poids 0.5)
    if abs(float(operation.montant) - float(ecriture.montant)) < 0.01:
        score += 0.5
    elif abs(float(operation.montant) - float(ecriture.montant)) / max(float(operation.montant), 1) < 0.05:
        score += 0.25  # Tolerance 5%

    # 2. Date +/- 3 jours (poids 0.3)
    delta = abs((operation.date_operation - ecriture.date_ecriture).days)
    if delta == 0:
        score += 0.3
    elif delta <= 3:
        score += 0.3 * (1 - delta / 4)

    # 3. Libelle Levenshtein (poids 0.2)
    sim = lev_ratio(operation.libelle.lower(), ecriture.libelle.lower())
    score += 0.2 * sim

    return round(score, 3)


def rapprocher_automatique(releve):
    """Lance le rapprochement automatique pour un releve"""
    from comptabilite.models import Ecriture
    from .models import OperationBancaire

    operations = OperationBancaire.objects.filter(releve=releve, rapprochee=False)
    ecritures = Ecriture.objects.filter(entreprise=releve.entreprise).exclude(
        operations_rapprochees__rapprochee=True
    )

    nb_auto = 0
    for op in operations:
        best_score = 0.0
        best_ec = None
        for ec in ecritures:
            score = calculate_score(op, ec)
            if score > best_score:
                best_score = score
                best_ec = ec

        if best_ec and best_score >= 0.85:
            op.ecriture_associee = best_ec
            op.score_rapprochement = best_score
            op.rapprochee = True
            op.save()
            nb_auto += 1
        elif best_ec and best_score >= 0.60:
            op.ecriture_associee = best_ec
            op.score_rapprochement = best_score
            op.suggeree = True
            op.save()

    releve.nb_rapprochees = nb_auto
    releve.save()
    return nb_auto
'''

FILES["rapprochement/views.py"] = '''import csv, io
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse
from datetime import datetime
from .models import ReleveBancaire, OperationBancaire
from .matcher import rapprocher_automatique

@login_required
def rapprochement_page(request):
    if not hasattr(request.user, 'entreprise'):
        return redirect('dashboard')
    releves = ReleveBancaire.objects.filter(entreprise=request.user.entreprise)
    return render(request, 'rapprochement/index.html', {'releves': releves})

@login_required
def import_releve(request):
    if request.method != 'POST' or not hasattr(request.user, 'entreprise'):
        return redirect('rapprochement_page')

    fichier = request.FILES.get('fichier')
    if not fichier:
        messages.error(request, "Fichier requis")
        return redirect('rapprochement_page')

    try:
        releve = ReleveBancaire.objects.create(
            entreprise=request.user.entreprise,
            nom=fichier.name,
            banque=request.POST.get('banque', 'Banque'),
        )

        # Parser CSV (format: date;libelle;montant;sens)
        decoded = fichier.read().decode('utf-8-sig')
        reader = csv.DictReader(io.StringIO(decoded), delimiter=';')

        nb = 0
        for row in reader:
            try:
                date_str = row.get('date', '').strip()
                date_op = None
                for fmt in ['%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y']:
                    try:
                        date_op = datetime.strptime(date_str, fmt).date()
                        break
                    except: pass
                if not date_op: continue

                OperationBancaire.objects.create(
                    releve=releve,
                    entreprise=request.user.entreprise,
                    date_operation=date_op,
                    libelle=row.get('libelle', '').strip()[:300],
                    montant=float(row.get('montant', '0').replace(',', '.').replace(' ', '')),
                    sens=row.get('sens', 'DEBIT').upper(),
                )
                nb += 1
            except Exception as e:
                print(f"Erreur ligne: {e}")
                continue

        releve.nb_operations = nb
        releve.save()

        # Lancer rapprochement automatique
        nb_auto = rapprocher_automatique(releve)
        messages.success(request, f"{nb} operations importees. {nb_auto} rapprochees automatiquement.")

    except Exception as e:
        messages.error(request, f"Erreur: {e}")

    return redirect('rapprochement_page')

@login_required
def detail_releve(request, pk):
    releve = get_object_or_404(ReleveBancaire, pk=pk, entreprise=request.user.entreprise)
    operations = releve.operations.all()
    return render(request, 'rapprochement/detail.html', {'releve': releve, 'operations': operations})

@login_required
def valider_rapprochement(request, op_id):
    op = get_object_or_404(OperationBancaire, pk=op_id, entreprise=request.user.entreprise)
    op.rapprochee = True
    op.suggeree = False
    op.save()
    return JsonResponse({'status': 'ok'})

@login_required
def telecharger_modele(request):
    """Telecharge un modele CSV"""
    from django.http import HttpResponse
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = 'attachment; filename="modele_releve_bancaire.csv"'
    response.write('\\ufeff')
    response.write('date;libelle;montant;sens\\n')
    response.write('15/06/2026;Virement client ABC;500000;CREDIT\\n')
    response.write('16/06/2026;Loyer juin 2026;280000;DEBIT\\n')
    response.write('17/06/2026;Facture CIE;68000;DEBIT\\n')
    return response
'''

FILES["rapprochement/urls.py"] = '''from django.urls import path
from . import views
urlpatterns = [
    path('', views.rapprochement_page, name='rapprochement_page'),
    path('import/', views.import_releve, name='import_releve'),
    path('releve/<int:pk>/', views.detail_releve, name='detail_releve'),
    path('valider/<int:op_id>/', views.valider_rapprochement, name='valider_rapprochement'),
    path('modele-csv/', views.telecharger_modele, name='modele_csv'),
]
'''

FILES["rapprochement/admin.py"] = '''from django.contrib import admin
from .models import ReleveBancaire, OperationBancaire

@admin.register(ReleveBancaire)
class ReleveBancaireAdmin(admin.ModelAdmin):
    list_display = ('nom', 'banque', 'entreprise', 'nb_operations', 'nb_rapprochees', 'date_import')

@admin.register(OperationBancaire)
class OperationBancaireAdmin(admin.ModelAdmin):
    list_display = ('date_operation', 'libelle', 'montant', 'sens', 'rapprochee', 'score_rapprochement')
    list_filter = ('rapprochee', 'sens')
'''

# ============================================
# REPORTING FINANCIER
# ============================================
FILES["reporting/__init__.py"] = ""
FILES["reporting/migrations/__init__.py"] = ""

FILES["reporting/apps.py"] = '''from django.apps import AppConfig
class ReportingConfig(AppConfig):
    default_auto_field = 'django.db.models.BigAutoField'
    name = 'reporting'
'''

FILES["reporting/models.py"] = '''# Pas de modeles, on calcule a la volee depuis les ecritures
'''

FILES["reporting/calculator.py"] = '''"""
Calcul des etats financiers SYSCOHADA
"""
from django.db.models import Sum, Q
from comptabilite.models import Ecriture

def calcul_bilan(entreprise, exercice=None):
    """Calcule le bilan (Actif/Passif) selon SYSCOHADA"""
    ecritures = Ecriture.objects.filter(entreprise=entreprise)
    if exercice:
        ecritures = ecritures.filter(date_ecriture__year=exercice)

    # ACTIF
    actif = {
        'immobilisations': {
            'terrains': solde_compte(ecritures, '211'),
            'batiments': solde_compte(ecritures, '213'),
            'materiel': solde_compte(ecritures, '244'),
            'transport': solde_compte(ecritures, '245'),
        },
        'creances': {
            'clients': solde_compte(ecritures, '411'),
        },
        'tresorerie': {
            'banque': solde_compte(ecritures, '521'),
            'caisse': solde_compte(ecritures, '531'),
        },
    }

    # PASSIF
    passif = {
        'capitaux': {
            'capital': solde_compte(ecritures, '101', sens='credit'),
            'reserves': solde_compte(ecritures, '106', sens='credit'),
        },
        'dettes': {
            'fournisseurs': solde_compte(ecritures, '401', sens='credit'),
            'personnel': solde_compte(ecritures, '421', sens='credit'),
            'tva': solde_compte(ecritures, '443', sens='credit'),
        },
    }

    # Totaux
    total_actif = sum_dict(actif)
    total_passif = sum_dict(passif)

    return {
        'actif': actif,
        'passif': passif,
        'total_actif': total_actif,
        'total_passif': total_passif,
    }


def calcul_compte_resultat(entreprise, exercice=None):
    """Calcule le compte de resultat"""
    ecritures = Ecriture.objects.filter(entreprise=entreprise)
    if exercice:
        ecritures = ecritures.filter(date_ecriture__year=exercice)

    # CHARGES (Classe 6)
    charges = {
        '601_achats': solde_compte(ecritures, '601'),
        '605_autres_achats': solde_compte(ecritures, '605'),
        '622_locations': solde_compte(ecritures, '622'),
        '624_entretien': solde_compte(ecritures, '624'),
        '626_telecom': solde_compte(ecritures, '626'),
        '631_personnel': solde_compte(ecritures, '631'),
        '641_impots': solde_compte(ecritures, '641'),
    }
    total_charges = sum(charges.values())

    # PRODUITS (Classe 7)
    produits = {
        '701_ventes': solde_compte(ecritures, '701', sens='credit'),
        '706_services': solde_compte(ecritures, '706', sens='credit'),
        '707_accessoires': solde_compte(ecritures, '707', sens='credit'),
    }
    total_produits = sum(produits.values())

    resultat = total_produits - total_charges

    return {
        'charges': charges,
        'produits': produits,
        'total_charges': total_charges,
        'total_produits': total_produits,
        'resultat': resultat,
    }


def solde_compte(ecritures, numero, sens='debit'):
    """Calcule le solde d'un compte"""
    if sens == 'debit':
        return float(ecritures.filter(compte_debit__numero=numero).aggregate(s=Sum('montant'))['s'] or 0)
    else:
        return float(ecritures.filter(compte_credit__numero=numero).aggregate(s=Sum('montant'))['s'] or 0)


def sum_dict(d):
    total = 0
    for k, v in d.items():
        if isinstance(v, dict):
            total += sum_dict(v)
        else:
            total += v
    return total


def grand_livre(entreprise, exercice=None):
    """Retourne toutes les ecritures pour le grand livre"""
    ecritures = Ecriture.objects.filter(entreprise=entreprise).order_by('date_ecriture', 'id')
    if exercice:
        ecritures = ecritures.filter(date_ecriture__year=exercice)
    return ecritures
'''

FILES["reporting/exports.py"] = '''"""
Export PDF et Excel des etats financiers
"""
from io import BytesIO
from django.http import HttpResponse

def export_bilan_pdf(bilan, entreprise):
    """Genere un PDF du bilan"""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=40)
    styles = getSampleStyleSheet()
    elements = []

    # Titre
    elements.append(Paragraph(f"<b>BILAN SYSCOHADA</b>", styles['Title']))
    elements.append(Paragraph(f"{entreprise.nom} - Exercice {entreprise.exercice_courant}", styles['Heading2']))
    elements.append(Spacer(1, 20))

    # ACTIF
    elements.append(Paragraph("<b>ACTIF</b>", styles['Heading2']))
    actif_data = [['Poste', 'Montant (FCFA)']]
    actif_data.append(['Immobilisations corporelles', ''])
    for k, v in bilan['actif']['immobilisations'].items():
        actif_data.append([f"  {k.capitalize()}", f"{v:,.0f}"])
    actif_data.append(['Creances', ''])
    for k, v in bilan['actif']['creances'].items():
        actif_data.append([f"  {k.capitalize()}", f"{v:,.0f}"])
    actif_data.append(['Tresorerie', ''])
    for k, v in bilan['actif']['tresorerie'].items():
        actif_data.append([f"  {k.capitalize()}", f"{v:,.0f}"])
    actif_data.append(['TOTAL ACTIF', f"{bilan['total_actif']:,.0f}"])

    t = Table(actif_data, colWidths=[300, 150])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#14706b')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor('#f4b860')),
        ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('PADDING', (0,0), (-1,-1), 6),
    ]))
    elements.append(t)
    elements.append(Spacer(1, 20))

    # PASSIF
    elements.append(Paragraph("<b>PASSIF</b>", styles['Heading2']))
    passif_data = [['Poste', 'Montant (FCFA)']]
    passif_data.append(['Capitaux propres', ''])
    for k, v in bilan['passif']['capitaux'].items():
        passif_data.append([f"  {k.capitalize()}", f"{v:,.0f}"])
    passif_data.append(['Dettes', ''])
    for k, v in bilan['passif']['dettes'].items():
        passif_data.append([f"  {k.capitalize()}", f"{v:,.0f}"])
    passif_data.append(['TOTAL PASSIF', f"{bilan['total_passif']:,.0f}"])

    t = Table(passif_data, colWidths=[300, 150])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#14706b')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor('#f4b860')),
        ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('PADDING', (0,0), (-1,-1), 6),
    ]))
    elements.append(t)

    doc.build(elements)
    buffer.seek(0)
    return buffer


def export_grand_livre_excel(ecritures, entreprise):
    """Export Excel du grand livre"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Grand Livre"

    # Header
    headers = ['Date', 'Piece', 'Libelle', 'Compte Debit', 'Compte Credit', 'Montant (FCFA)', 'Statut']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='14706B')
        cell.alignment = Alignment(horizontal='center')

    # Donnees
    for e in ecritures:
        ws.append([
            e.date_ecriture.strftime('%d/%m/%Y'),
            e.numero_piece or '-',
            e.libelle,
            f"{e.compte_debit.numero} - {e.compte_debit.libelle}",
            f"{e.compte_credit.numero} - {e.compte_credit.libelle}",
            float(e.montant),
            e.get_statut_display(),
        ])

    # Largeurs colonnes
    for col, width in enumerate([12, 12, 40, 30, 30, 18, 12], start=1):
        ws.column_dimensions[chr(64+col)].width = width

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
'''

FILES["reporting/views.py"] = '''from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from .calculator import calcul_bilan, calcul_compte_resultat, grand_livre
from .exports import export_bilan_pdf, export_grand_livre_excel

@login_required
def reporting_index(request):
    if not hasattr(request.user, 'entreprise'):
        return redirect('dashboard')
    return render(request, 'reporting/index.html', {})

@login_required
def voir_bilan(request):
    if not hasattr(request.user, 'entreprise'):
        return redirect('dashboard')
    bilan = calcul_bilan(request.user.entreprise)
    return render(request, 'reporting/bilan.html', {'bilan': bilan})

@login_required
def voir_compte_resultat(request):
    if not hasattr(request.user, 'entreprise'):
        return redirect('dashboard')
    cr = calcul_compte_resultat(request.user.entreprise)
    return render(request, 'reporting/compte_resultat.html', {'cr': cr})

@login_required
def voir_grand_livre(request):
    if not hasattr(request.user, 'entreprise'):
        return redirect('dashboard')
    ecritures = grand_livre(request.user.entreprise)
    return render(request, 'reporting/grand_livre.html', {'ecritures': ecritures})

@login_required
def export_bilan(request):
    bilan = calcul_bilan(request.user.entreprise)
    buffer = export_bilan_pdf(bilan, request.user.entreprise)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="bilan_{request.user.entreprise.nom}.pdf"'
    return response

@login_required
def export_grand_livre(request):
    ecritures = grand_livre(request.user.entreprise)
    buffer = export_grand_livre_excel(ecritures, request.user.entreprise)
    response = HttpResponse(buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="grand_livre_{request.user.entreprise.nom}.xlsx"'
    return response
'''

FILES["reporting/urls.py"] = '''from django.urls import path
from . import views
urlpatterns = [
    path('', views.reporting_index, name='reporting_index'),
    path('bilan/', views.voir_bilan, name='voir_bilan'),
    path('compte-resultat/', views.voir_compte_resultat, name='voir_compte_resultat'),
    path('grand-livre/', views.voir_grand_livre, name='voir_grand_livre'),
    path('bilan/pdf/', views.export_bilan, name='export_bilan_pdf'),
    path('grand-livre/excel/', views.export_grand_livre, name='export_grand_livre_excel'),
]
'''

# ============================================
# ALERTES
# ============================================
FILES["alertes/__init__.py"] = ""
FILES["alertes/migrations/__init__.py"] = ""

FILES["alertes/apps.py"] = '''from django.apps import AppConfig
class AlertesConfig(AppConfig):
    default_auto_field = 'django.db.models.BigAutoField'
    name = 'alertes'
'''

FILES["alertes/models.py"] = '''from django.db import models
from accounts.models import Entreprise

class SeuilAlerte(models.Model):
    TYPES = [
        ('TRESORERIE_BASSE', 'Tresorerie basse'),
        ('CREANCE_ELEVEE', 'Creances elevees'),
        ('DETTE_ELEVEE', 'Dettes elevees'),
        ('IMPAYE', 'Facture impayee'),
    ]
    entreprise = models.ForeignKey(Entreprise, on_delete=models.CASCADE, related_name='seuils')
    type_alerte = models.CharField(max_length=30, choices=TYPES)
    seuil = models.DecimalField(max_digits=15, decimal_places=2)
    actif = models.BooleanField(default=True)
    notif_email = models.BooleanField(default=True)
    class Meta:
        db_table = 'seuils_alertes'

class Alerte(models.Model):
    NIVEAUX = [('INFO', 'Info'), ('WARNING', 'Avertissement'), ('DANGER', 'Critique')]
    entreprise = models.ForeignKey(Entreprise, on_delete=models.CASCADE, related_name='alertes')
    niveau = models.CharField(max_length=20, choices=NIVEAUX, default='WARNING')
    titre = models.CharField(max_length=200)
    message = models.TextField()
    lue = models.BooleanField(default=False)
    created_at = models.DateTimeField(auto_now_add=True)
    class Meta:
        db_table = 'alertes'
        ordering = ['-created_at']
'''

FILES["alertes/checker.py"] = '''"""
Verificateur d'alertes - lance les checks automatiquement
"""
from django.db.models import Sum
from comptabilite.models import Ecriture
from .models import Alerte, SeuilAlerte

def check_alertes(entreprise):
    """Verifie les seuils et cree les alertes"""
    nb_alertes = 0
    seuils = SeuilAlerte.objects.filter(entreprise=entreprise, actif=True)

    for seuil in seuils:
        if seuil.type_alerte == 'TRESORERIE_BASSE':
            tresorerie = Ecriture.objects.filter(
                entreprise=entreprise,
                compte_debit__numero__in=['521', '531']
            ).aggregate(s=Sum('montant'))['s'] or 0

            if tresorerie < seuil.seuil:
                if not Alerte.objects.filter(
                    entreprise=entreprise, titre='Tresorerie basse', lue=False
                ).exists():
                    Alerte.objects.create(
                        entreprise=entreprise, niveau='DANGER',
                        titre='Tresorerie basse',
                        message=f'Tresorerie actuelle : {tresorerie:,.0f} FCFA (seuil: {seuil.seuil:,.0f} FCFA)'
                    )
                    nb_alertes += 1

        elif seuil.type_alerte == 'CREANCE_ELEVEE':
            creances = Ecriture.objects.filter(
                entreprise=entreprise, compte_debit__numero='411'
            ).aggregate(s=Sum('montant'))['s'] or 0

            if creances > seuil.seuil:
                if not Alerte.objects.filter(
                    entreprise=entreprise, titre='Creances elevees', lue=False
                ).exists():
                    Alerte.objects.create(
                        entreprise=entreprise, niveau='WARNING',
                        titre='Creances elevees',
                        message=f'Creances clients : {creances:,.0f} FCFA (seuil: {seuil.seuil:,.0f} FCFA)'
                    )
                    nb_alertes += 1

    return nb_alertes


def init_seuils_par_defaut(entreprise):
    """Initialise les seuils par defaut"""
    defaults = [
        ('TRESORERIE_BASSE', 500000),
        ('CREANCE_ELEVEE', 2000000),
        ('DETTE_ELEVEE', 1000000),
    ]
    for type_a, seuil in defaults:
        SeuilAlerte.objects.get_or_create(
            entreprise=entreprise, type_alerte=type_a,
            defaults={'seuil': seuil}
        )
'''

FILES["alertes/views.py"] = '''from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse
from .models import SeuilAlerte, Alerte
from .checker import check_alertes, init_seuils_par_defaut

@login_required
def alertes_page(request):
    if not hasattr(request.user, 'entreprise'):
        return redirect('dashboard')

    # Init seuils par defaut si vide
    if not SeuilAlerte.objects.filter(entreprise=request.user.entreprise).exists():
        init_seuils_par_defaut(request.user.entreprise)

    # Lancer check
    check_alertes(request.user.entreprise)

    seuils = SeuilAlerte.objects.filter(entreprise=request.user.entreprise)
    alertes = Alerte.objects.filter(entreprise=request.user.entreprise)
    return render(request, 'alertes/index.html', {'seuils': seuils, 'alertes': alertes})

@login_required
def modifier_seuil(request, pk):
    seuil = get_object_or_404(SeuilAlerte, pk=pk, entreprise=request.user.entreprise)
    if request.method == 'POST':
        seuil.seuil = request.POST.get('seuil', seuil.seuil)
        seuil.actif = request.POST.get('actif') == 'on'
        seuil.notif_email = request.POST.get('notif_email') == 'on'
        seuil.save()
        messages.success(request, "Seuil modifie")
    return redirect('alertes_page')

@login_required
def marquer_lue(request, pk):
    a = get_object_or_404(Alerte, pk=pk, entreprise=request.user.entreprise)
    a.lue = True
    a.save()
    return JsonResponse({'status': 'ok'})
'''

FILES["alertes/urls.py"] = '''from django.urls import path
from . import views
urlpatterns = [
    path('', views.alertes_page, name='alertes_page'),
    path('seuil/<int:pk>/', views.modifier_seuil, name='modifier_seuil'),
    path('lue/<int:pk>/', views.marquer_lue, name='marquer_lue'),
]
'''

FILES["alertes/admin.py"] = '''from django.contrib import admin
from .models import SeuilAlerte, Alerte

@admin.register(SeuilAlerte)
class SeuilAlerteAdmin(admin.ModelAdmin):
    list_display = ('entreprise', 'type_alerte', 'seuil', 'actif')

@admin.register(Alerte)
class AlerteAdmin(admin.ModelAdmin):
    list_display = ('titre', 'niveau', 'entreprise', 'lue', 'created_at')
    list_filter = ('niveau', 'lue')
'''

# ============================================
# TEMPLATES RAPPROCHEMENT
# ============================================
FILES["templates/rapprochement/index.html"] = '''{% extends 'base.html' %}{% block page_title %}Rapprochement bancaire{% endblock %}{% block page_subtitle %}Score = Montant (0.5) + Date +/-3j (0.3) + Libelle Levenshtein (0.2){% endblock %}{% block content %}
<div class="app-layout">{% include 'partials/sidebar.html' %}<main class="main-content">{% include 'partials/topbar.html' %}
<div class="content-area">
{% if messages %}{% for m in messages %}<div class="alert alert-{{ m.tags }} alert-dismissible fade show">{{ m }}<button type="button" class="btn-close" data-bs-dismiss="alert"></button></div>{% endfor %}{% endif %}

<div class="row g-4">
<div class="col-md-5"><div class="card"><div class="card-body">
<h4>Importer un releve bancaire</h4>
<p class="text-muted">Format CSV (date;libelle;montant;sens)</p>
<a href="{% url 'modele_csv' %}" class="btn btn-outline-secondary btn-sm mb-3"><i class="bi bi-download"></i> Telecharger le modele</a>
<form method="POST" action="{% url 'import_releve' %}" enctype="multipart/form-data">{% csrf_token %}
<div class="mb-3"><label>Banque</label><input type="text" name="banque" class="form-control" placeholder="SGBCI, BICICI..." required></div>
<div class="mb-3"><label>Fichier CSV</label><input type="file" name="fichier" class="form-control" accept=".csv" required></div>
<button type="submit" class="btn btn-connect w-100"><i class="bi bi-upload"></i> Importer & Rapprocher</button>
</form>
</div></div></div>

<div class="col-md-7"><div class="card"><div class="card-body">
<h4>Releves importes</h4>
{% for r in releves %}
<a href="{% url 'detail_releve' r.pk %}" class="d-block text-decoration-none">
<div class="facture-card">
<i class="bi bi-bank"></i>
<div class="flex-grow-1">
<strong>{{ r.banque }} - {{ r.nom }}</strong><br>
<small>{{ r.date_import|date:"d/m/Y H:i" }} - {{ r.nb_operations }} operations - <span class="badge bg-success">{{ r.nb_rapprochees }} rapprochees</span></small>
</div>
</div></a>
{% empty %}
<p class="text-muted">Aucun releve importe.</p>
{% endfor %}
</div></div></div>
</div>

</div></main></div>
{% endblock %}
'''

FILES["templates/rapprochement/detail.html"] = '''{% extends 'base.html' %}{% block page_title %}{{ releve.banque }} - {{ releve.nom }}{% endblock %}{% block content %}
<div class="app-layout">{% include 'partials/sidebar.html' %}<main class="main-content">{% include 'partials/topbar.html' %}
<div class="content-area">
<a href="{% url 'rapprochement_page' %}" class="btn btn-link"><i class="bi bi-arrow-left"></i> Retour</a>
<div class="card mt-3"><div class="card-body">
<h4>{{ releve.nb_operations }} operations - {{ releve.nb_rapprochees }} rapprochees automatiquement</h4>
<table class="table">
<thead><tr><th>Date</th><th>Libelle</th><th>Montant</th><th>Sens</th><th>Score</th><th>Statut</th><th>Ecriture associee</th></tr></thead>
<tbody>
{% for op in operations %}
<tr>
<td>{{ op.date_operation|date:"d/m/Y" }}</td>
<td>{{ op.libelle }}</td>
<td>{{ op.montant }} F</td>
<td><span class="badge bg-{% if op.sens == 'CREDIT' %}success{% else %}danger{% endif %}">{{ op.sens }}</span></td>
<td>{% if op.score_rapprochement %}<strong>{{ op.score_rapprochement|floatformat:2 }}</strong>{% else %}-{% endif %}</td>
<td>{% if op.rapprochee %}<span class="badge bg-success">Rapprochee</span>{% elif op.suggeree %}<span class="badge bg-warning">Suggeree</span>{% else %}<span class="badge bg-secondary">Non rapprochee</span>{% endif %}</td>
<td>{% if op.ecriture_associee %}{{ op.ecriture_associee.libelle }}{% if op.suggeree %}<button class="btn btn-sm btn-success ms-2" onclick="valider({{ op.id }})"><i class="bi bi-check"></i> Valider</button>{% endif %}{% else %}-{% endif %}</td>
</tr>
{% endfor %}
</tbody>
</table>
</div></div>
</div></main></div>
<script>
function getCookie(n){const v=`; ${document.cookie}`.split(`; ${n}=`);if(v.length===2)return v.pop().split(';').shift();}
async function valider(id){const r=await fetch('/rapprochement/valider/'+id+'/',{method:'POST',headers:{'X-CSRFToken':getCookie('csrftoken')}});if(r.ok)location.reload();}
</script>
{% endblock %}
'''

# ============================================
# TEMPLATES REPORTING
# ============================================
FILES["templates/reporting/index.html"] = '''{% extends 'base.html' %}{% block page_title %}Reporting financier{% endblock %}{% block page_subtitle %}Etats financiers SYSCOHADA{% endblock %}{% block content %}
<div class="app-layout">{% include 'partials/sidebar.html' %}<main class="main-content">{% include 'partials/topbar.html' %}
<div class="content-area">
<div class="row g-4">
<div class="col-md-4"><div class="card"><div class="card-body text-center">
<i class="bi bi-file-earmark-bar-graph" style="font-size:3rem;color:#14706b;"></i>
<h5 class="mt-3">Bilan SYSCOHADA</h5>
<p class="text-muted">Actif / Passif</p>
<a href="{% url 'voir_bilan' %}" class="btn btn-connect w-100 mb-2">Voir</a>
<a href="{% url 'export_bilan_pdf' %}" class="btn btn-outline-primary w-100"><i class="bi bi-file-pdf"></i> Export PDF</a>
</div></div></div>

<div class="col-md-4"><div class="card"><div class="card-body text-center">
<i class="bi bi-graph-up-arrow" style="font-size:3rem;color:#14706b;"></i>
<h5 class="mt-3">Compte de resultat</h5>
<p class="text-muted">Charges / Produits</p>
<a href="{% url 'voir_compte_resultat' %}" class="btn btn-connect w-100">Voir</a>
</div></div></div>

<div class="col-md-4"><div class="card"><div class="card-body text-center">
<i class="bi bi-journal-bookmark" style="font-size:3rem;color:#14706b;"></i>
<h5 class="mt-3">Grand Livre</h5>
<p class="text-muted">Toutes les ecritures</p>
<a href="{% url 'voir_grand_livre' %}" class="btn btn-connect w-100 mb-2">Voir</a>
<a href="{% url 'export_grand_livre_excel' %}" class="btn btn-outline-primary w-100"><i class="bi bi-file-excel"></i> Export Excel</a>
</div></div></div>
</div>
</div></main></div>
{% endblock %}
'''

FILES["templates/reporting/bilan.html"] = '''{% extends 'base.html' %}{% block page_title %}Bilan SYSCOHADA{% endblock %}{% block content %}
<div class="app-layout">{% include 'partials/sidebar.html' %}<main class="main-content">{% include 'partials/topbar.html' %}
<div class="content-area">
<a href="{% url 'reporting_index' %}" class="btn btn-link"><i class="bi bi-arrow-left"></i> Retour</a>
<a href="{% url 'export_bilan_pdf' %}" class="btn btn-outline-primary float-end"><i class="bi bi-file-pdf"></i> Export PDF</a>

<div class="row g-4 mt-3">
<div class="col-md-6"><div class="card"><div class="card-body">
<h4 style="color:#14706b;">ACTIF</h4>
<table class="table">
<tr><td><strong>Immobilisations</strong></td><td></td></tr>
{% for k, v in bilan.actif.immobilisations.items %}<tr><td>&nbsp;&nbsp;{{ k|capfirst }}</td><td class="text-end">{{ v|floatformat:0 }} F</td></tr>{% endfor %}
<tr><td><strong>Creances</strong></td><td></td></tr>
{% for k, v in bilan.actif.creances.items %}<tr><td>&nbsp;&nbsp;{{ k|capfirst }}</td><td class="text-end">{{ v|floatformat:0 }} F</td></tr>{% endfor %}
<tr><td><strong>Tresorerie</strong></td><td></td></tr>
{% for k, v in bilan.actif.tresorerie.items %}<tr><td>&nbsp;&nbsp;{{ k|capfirst }}</td><td class="text-end">{{ v|floatformat:0 }} F</td></tr>{% endfor %}
<tr style="background:#f4b860;"><td><strong>TOTAL ACTIF</strong></td><td class="text-end"><strong>{{ bilan.total_actif|floatformat:0 }} F</strong></td></tr>
</table>
</div></div></div>

<div class="col-md-6"><div class="card"><div class="card-body">
<h4 style="color:#14706b;">PASSIF</h4>
<table class="table">
<tr><td><strong>Capitaux propres</strong></td><td></td></tr>
{% for k, v in bilan.passif.capitaux.items %}<tr><td>&nbsp;&nbsp;{{ k|capfirst }}</td><td class="text-end">{{ v|floatformat:0 }} F</td></tr>{% endfor %}
<tr><td><strong>Dettes</strong></td><td></td></tr>
{% for k, v in bilan.passif.dettes.items %}<tr><td>&nbsp;&nbsp;{{ k|capfirst }}</td><td class="text-end">{{ v|floatformat:0 }} F</td></tr>{% endfor %}
<tr style="background:#f4b860;"><td><strong>TOTAL PASSIF</strong></td><td class="text-end"><strong>{{ bilan.total_passif|floatformat:0 }} F</strong></td></tr>
</table>
</div></div></div>
</div>

</div></main></div>
{% endblock %}
'''

FILES["templates/reporting/compte_resultat.html"] = '''{% extends 'base.html' %}{% block page_title %}Compte de resultat{% endblock %}{% block content %}
<div class="app-layout">{% include 'partials/sidebar.html' %}<main class="main-content">{% include 'partials/topbar.html' %}
<div class="content-area">
<a href="{% url 'reporting_index' %}" class="btn btn-link"><i class="bi bi-arrow-left"></i> Retour</a>
<div class="row g-4 mt-3">
<div class="col-md-6"><div class="card"><div class="card-body">
<h4 style="color:#c0392b;">CHARGES</h4>
<table class="table">
{% for k, v in cr.charges.items %}<tr><td>{{ k }}</td><td class="text-end">{{ v|floatformat:0 }} F</td></tr>{% endfor %}
<tr style="background:#fde0e0;"><td><strong>TOTAL CHARGES</strong></td><td class="text-end"><strong>{{ cr.total_charges|floatformat:0 }} F</strong></td></tr>
</table>
</div></div></div>
<div class="col-md-6"><div class="card"><div class="card-body">
<h4 style="color:#14706b;">PRODUITS</h4>
<table class="table">
{% for k, v in cr.produits.items %}<tr><td>{{ k }}</td><td class="text-end">{{ v|floatformat:0 }} F</td></tr>{% endfor %}
<tr style="background:#d4ede9;"><td><strong>TOTAL PRODUITS</strong></td><td class="text-end"><strong>{{ cr.total_produits|floatformat:0 }} F</strong></td></tr>
</table>
</div></div></div>
</div>
<div class="card mt-4" style="background:{% if cr.resultat >= 0 %}#d4ede9{% else %}#fde0e0{% endif %};"><div class="card-body text-center">
<h2>RESULTAT NET : <strong>{{ cr.resultat|floatformat:0 }} FCFA</strong></h2>
<p>{% if cr.resultat >= 0 %}Benefice{% else %}Perte{% endif %}</p>
</div></div>
</div></main></div>
{% endblock %}
'''

FILES["templates/reporting/grand_livre.html"] = '''{% extends 'base.html' %}{% block page_title %}Grand Livre{% endblock %}{% block content %}
<div class="app-layout">{% include 'partials/sidebar.html' %}<main class="main-content">{% include 'partials/topbar.html' %}
<div class="content-area">
<a href="{% url 'reporting_index' %}" class="btn btn-link"><i class="bi bi-arrow-left"></i> Retour</a>
<a href="{% url 'export_grand_livre_excel' %}" class="btn btn-outline-primary float-end"><i class="bi bi-file-excel"></i> Export Excel</a>
<div class="card mt-3"><div class="card-body">
<table class="table">
<thead><tr><th>Date</th><th>Piece</th><th>Libelle</th><th>Debit</th><th>Credit</th><th>Montant</th><th>Statut</th></tr></thead>
<tbody>
{% for e in ecritures %}
<tr><td>{{ e.date_ecriture|date:"d/m/Y" }}</td><td>{{ e.numero_piece|default:"-" }}</td><td>{{ e.libelle }}</td><td>{{ e.compte_debit.numero }}</td><td>{{ e.compte_credit.numero }}</td><td>{{ e.montant }} F</td><td><span class="badge bg-{% if e.statut == 'VALIDEE' %}success{% elif e.statut == 'AUTO' %}info{% else %}secondary{% endif %}">{{ e.get_statut_display }}</span></td></tr>
{% empty %}<tr><td colspan="7" class="text-center text-muted">Aucune ecriture</td></tr>{% endfor %}
</tbody></table>
</div></div>
</div></main></div>
{% endblock %}
'''

# ============================================
# TEMPLATE ALERTES
# ============================================
FILES["templates/alertes/index.html"] = '''{% extends 'base.html' %}{% block page_title %}Alertes & Seuils{% endblock %}{% block page_subtitle %}Seuils configurables (tresorerie, impayes, echeances){% endblock %}{% block content %}
<div class="app-layout">{% include 'partials/sidebar.html' %}<main class="main-content">{% include 'partials/topbar.html' %}
<div class="content-area">
{% if messages %}{% for m in messages %}<div class="alert alert-{{ m.tags }} alert-dismissible fade show">{{ m }}<button type="button" class="btn-close" data-bs-dismiss="alert"></button></div>{% endfor %}{% endif %}

<div class="row g-4">
<div class="col-md-6"><div class="card"><div class="card-body">
<h4>Seuils configurables</h4>
{% for s in seuils %}
<form method="POST" action="{% url 'modifier_seuil' s.pk %}" class="border-bottom py-3">{% csrf_token %}
<div class="d-flex justify-content-between align-items-center">
<div><strong>{{ s.get_type_alerte_display }}</strong></div>
<div><span class="badge bg-{% if s.actif %}success{% else %}secondary{% endif %}">{% if s.actif %}Actif{% else %}Inactif{% endif %}</span></div>
</div>
<div class="row mt-2">
<div class="col-md-6"><input type="number" name="seuil" class="form-control" value="{{ s.seuil|floatformat:0 }}" step="1000"></div>
<div class="col-md-3"><label><input type="checkbox" name="actif" {% if s.actif %}checked{% endif %}> Actif</label></div>
<div class="col-md-3"><button type="submit" class="btn btn-sm btn-connect">Sauver</button></div>
</div>
</form>
{% endfor %}
</div></div></div>

<div class="col-md-6"><div class="card"><div class="card-body">
<h4>Alertes recentes</h4>
{% for a in alertes %}
<div class="alert alert-{{ a.niveau|lower }} {% if a.lue %}opacity-50{% endif %}">
<div class="d-flex justify-content-between"><div><strong>{{ a.titre }}</strong></div><small>{{ a.created_at|date:"d/m H:i" }}</small></div>
<p class="mb-1">{{ a.message }}</p>
{% if not a.lue %}<button class="btn btn-sm btn-outline-secondary" onclick="marquerLue({{ a.pk }})">Marquer lue</button>{% endif %}
</div>
{% empty %}<p class="text-muted">Aucune alerte.</p>{% endfor %}
</div></div></div>
</div>

</div></main></div>
<script>
function getCookie(n){const v=`; ${document.cookie}`.split(`; ${n}=`);if(v.length===2)return v.pop().split(';').shift();}
async function marquerLue(id){const r=await fetch('/alertes/lue/'+id+'/',{method:'POST',headers:{'X-CSRFToken':getCookie('csrftoken')}});if(r.ok)location.reload();}
</script>
{% endblock %}
'''

# ============================================
# UPDATE SIDEBAR pour ajouter les nouveaux liens
# ============================================
FILES["templates/partials/sidebar.html"] = '''<aside class="sidebar">
<div class="sidebar-brand"><div class="logo-icon">C</div><div><h1 class="brand-title">ComptaAuto</h1><p class="brand-subtitle">COMPTABILITE</p></div></div>
<div class="sidebar-section">PILOTAGE</div>
<nav class="sidebar-nav"><a href="{% url 'dashboard' %}" class="nav-item {% if request.resolver_match.url_name == 'dashboard' %}active{% endif %}"><i class="bi bi-grid-1x2"></i> Tableau de bord</a></nav>
<div class="sidebar-section">OPERATIONS</div>
<nav class="sidebar-nav">
<a href="{% url 'ocr_page' %}" class="nav-item {% if 'ocr' in request.path %}active{% endif %}"><i class="bi bi-text-paragraph"></i> Saisie OCR</a>
<a href="{% url 'rapprochement_page' %}" class="nav-item {% if 'rapprochement' in request.path %}active{% endif %}"><i class="bi bi-arrow-left-right"></i> Rapprochement</a>
<a href="{% url 'ecritures' %}" class="nav-item {% if 'ecriture' in request.path %}active{% endif %}"><i class="bi bi-journal-text"></i> Ecritures</a>
<a href="{% url 'nouvelle_ecriture' %}" class="nav-item"><i class="bi bi-plus-circle"></i> Nouvelle ecriture</a>
</nav>
<div class="sidebar-section">REPORTING & ALERTES</div>
<nav class="sidebar-nav">
<a href="{% url 'reporting_index' %}" class="nav-item {% if 'reporting' in request.path %}active{% endif %}"><i class="bi bi-bar-chart-line"></i> Reporting financier</a>
<a href="{% url 'alertes_page' %}" class="nav-item {% if 'alerte' in request.path %}active{% endif %}"><i class="bi bi-bell"></i> Alertes</a>
</nav>
<div class="sidebar-section">CONFIGURATION</div>
<nav class="sidebar-nav">
<a href="{% url 'init_plan' %}" class="nav-item"><i class="bi bi-list-ol"></i> Initialiser plan</a>
</nav>
<div class="sidebar-footer"><strong>{{ user.entreprise.nom }}</strong><small>SYSCOHADA - Exercice {{ user.entreprise.exercice_courant }}</small></div>
</aside>
'''

# ============================================
# UPDATE settings.py + urls.py
# ============================================
FILES["comptaauto/settings.py"] = '''from pathlib import Path
import environ, os

BASE_DIR = Path(__file__).resolve().parent.parent
env = environ.Env(DEBUG=(bool, False))
environ.Env.read_env(os.path.join(BASE_DIR, '.env'))

SECRET_KEY = env('SECRET_KEY', default='dev-key-change-me')
DEBUG = env('DEBUG', default=True)
ALLOWED_HOSTS = env.list('ALLOWED_HOSTS', default=['localhost', '127.0.0.1', '*'])

INSTALLED_APPS = [
    'django.contrib.admin','django.contrib.auth','django.contrib.contenttypes',
    'django.contrib.sessions','django.contrib.messages','django.contrib.staticfiles',
    'accounts','comptabilite','ocr_app','dashboard',
    'rapprochement','reporting','alertes',
]

MIDDLEWARE = [
    'django.middleware.security.SecurityMiddleware',
    'whitenoise.middleware.WhiteNoiseMiddleware',
    'django.contrib.sessions.middleware.SessionMiddleware',
    'django.middleware.common.CommonMiddleware',
    'django.middleware.csrf.CsrfViewMiddleware',
    'django.contrib.auth.middleware.AuthenticationMiddleware',
    'django.contrib.messages.middleware.MessageMiddleware',
    'django.middleware.clickjacking.XFrameOptionsMiddleware',
]

ROOT_URLCONF = 'comptaauto.urls'
TEMPLATES = [{
    'BACKEND': 'django.template.backends.django.DjangoTemplates',
    'DIRS': [BASE_DIR / 'templates'],
    'APP_DIRS': True,
    'OPTIONS': {'context_processors': [
        'django.template.context_processors.debug',
        'django.template.context_processors.request',
        'django.contrib.auth.context_processors.auth',
        'django.contrib.messages.context_processors.messages',
    ]},
}]
WSGI_APPLICATION = 'comptaauto.wsgi.application'

DATABASES = {'default': {'ENGINE': 'django.db.backends.sqlite3', 'NAME': BASE_DIR / 'db.sqlite3'}}

AUTH_USER_MODEL = 'accounts.User'
AUTH_PASSWORD_VALIDATORS = [{'NAME': 'django.contrib.auth.password_validation.MinimumLengthValidator'}]

LANGUAGE_CODE = 'fr-fr'
TIME_ZONE = 'Africa/Abidjan'
USE_I18N = True
USE_TZ = True

STATIC_URL = '/static/'
STATIC_ROOT = BASE_DIR / 'staticfiles'
STATICFILES_DIRS = [BASE_DIR / 'static']
STATICFILES_STORAGE = 'whitenoise.storage.CompressedManifestStaticFilesStorage'
MEDIA_URL = '/media/'
MEDIA_ROOT = BASE_DIR / 'media'
DEFAULT_AUTO_FIELD = 'django.db.models.BigAutoField'
LOGIN_URL = '/login/'
LOGIN_REDIRECT_URL = '/dashboard/'
LOGOUT_REDIRECT_URL = '/login/'
OCR_SPACE_API_KEY = env('OCR_SPACE_API_KEY', default='helloworld')
'''

FILES["comptaauto/urls.py"] = '''from django.contrib import admin
from django.urls import path, include
from django.conf import settings
from django.conf.urls.static import static
from django.shortcuts import redirect

urlpatterns = [
    path('django-admin/', admin.site.urls),
    path('', lambda r: redirect('login')),
    path('', include('accounts.urls')),
    path('dashboard/', include('dashboard.urls')),
    path('admin-panel/', include('dashboard.urls_admin')),
    path('comptabilite/', include('comptabilite.urls')),
    path('ocr/', include('ocr_app.urls')),
    path('rapprochement/', include('rapprochement.urls')),
    path('reporting/', include('reporting.urls')),
    path('alertes/', include('alertes.urls')),
]

if settings.DEBUG:
    urlpatterns += static(settings.MEDIA_URL, document_root=settings.MEDIA_ROOT)
    urlpatterns += static(settings.STATIC_URL, document_root=settings.STATIC_ROOT)
'''

# Ecriture fichiers
for fp, content in FILES.items():
    p = BASE / fp
    p.parent.mkdir(parents=True, exist_ok=True)
    with open(p, 'w', encoding='utf-8') as f:
        f.write(content)
    print(f"      [OK] {fp}")

print(f"\n      {len(FILES)} fichiers crees/mis a jour\n")

# ============ MIGRATIONS ============
print("[4/4] Migrations base de donnees...")
subprocess.run([PY, "manage.py", "makemigrations", "rapprochement", "reporting", "alertes"], check=True)
subprocess.run([PY, "manage.py", "migrate"], check=True)
print("      OK")

print("\n" + "="*65)
print("  MODULES AJOUTES AVEC SUCCES !")
print("="*65)
print(f"\nRelancer le serveur :")
print(f"   {PY} manage.py runserver")
print(f"\nNouveaux modules disponibles dans la sidebar :")
print(f"  - Rapprochement bancaire  : /rapprochement/")
print(f"  - Reporting financier     : /reporting/")
print(f"  - Alertes configurables   : /alertes/")
print(f"\nTout est conforme a votre presentation !")
print("="*65 + "\n")
