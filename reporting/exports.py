"""
Export PDF et Excel des etats financiers
"""
from io import BytesIO
from django.http import HttpResponse

def export_bilan_pdf(bilan, entreprise):
    """Genere un PDF du bilan"""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=40)
    styles = getSampleStyleSheet()
    elements = []

    # Titre
    elements.append(Paragraph(f"<b>BILAN SYSCOHADA</b>", styles['Title']))
    elements.append(Paragraph(f"{entreprise.nom} - Exercice {entreprise.exercice_courant}", styles['Heading2']))
    elements.append(Spacer(1, 20))

    # ACTIF
    elements.append(Paragraph("<b>ACTIF</b>", styles['Heading2']))
    actif_data = [['Poste', 'Montant (FCFA)']]
    actif_data.append(['Immobilisations corporelles', ''])
    for k, v in bilan['actif']['immobilisations'].items():
        actif_data.append([f"  {k.capitalize()}", f"{v:,.0f}"])
    actif_data.append(['Creances', ''])
    for k, v in bilan['actif']['creances'].items():
        actif_data.append([f"  {k.capitalize()}", f"{v:,.0f}"])
    actif_data.append(['Tresorerie', ''])
    for k, v in bilan['actif']['tresorerie'].items():
        actif_data.append([f"  {k.capitalize()}", f"{v:,.0f}"])
    actif_data.append(['TOTAL ACTIF', f"{bilan['total_actif']:,.0f}"])

    t = Table(actif_data, colWidths=[300, 150])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#14706b')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor('#f4b860')),
        ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('PADDING', (0,0), (-1,-1), 6),
    ]))
    elements.append(t)
    elements.append(Spacer(1, 20))

    # PASSIF
    elements.append(Paragraph("<b>PASSIF</b>", styles['Heading2']))
    passif_data = [['Poste', 'Montant (FCFA)']]
    passif_data.append(['Capitaux propres', ''])
    for k, v in bilan['passif']['capitaux'].items():
        passif_data.append([f"  {k.capitalize()}", f"{v:,.0f}"])
    passif_data.append(['Dettes', ''])
    for k, v in bilan['passif']['dettes'].items():
        passif_data.append([f"  {k.capitalize()}", f"{v:,.0f}"])
    passif_data.append(['TOTAL PASSIF', f"{bilan['total_passif']:,.0f}"])

    t = Table(passif_data, colWidths=[300, 150])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#14706b')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor('#f4b860')),
        ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('PADDING', (0,0), (-1,-1), 6),
    ]))
    elements.append(t)

    doc.build(elements)
    buffer.seek(0)
    return buffer


def export_grand_livre_excel(ecritures, entreprise):
    """Export Excel du grand livre"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Grand Livre"

    # Header
    headers = ['Date', 'Piece', 'Libelle', 'Compte Debit', 'Compte Credit', 'Montant (FCFA)', 'Statut']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='14706B')
        cell.alignment = Alignment(horizontal='center')

    # Donnees
    for e in ecritures:
        ws.append([
            e.date_ecriture.strftime('%d/%m/%Y'),
            e.numero_piece or '-',
            e.libelle,
            f"{e.compte_debit.numero} - {e.compte_debit.libelle}",
            f"{e.compte_credit.numero} - {e.compte_credit.libelle}",
            float(e.montant),
            e.get_statut_display(),
        ])

    # Largeurs colonnes
    for col, width in enumerate([12, 12, 40, 30, 30, 18, 12], start=1):
        ws.column_dimensions[chr(64+col)].width = width

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
